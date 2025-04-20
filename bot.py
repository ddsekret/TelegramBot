import os
import re
import logging
import asyncio
import yadisk
import traceback
import sys
import requests
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command, StateFilter
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.client.session.aiohttp import AiohttpSession
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar  # Добавляем для работы с итогами месяца
from dotenv import load_dotenv
import pandas as pd  # Для работы с базами данных

# Настройка логирования с более подробным форматом
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s",
    handlers=[
        logging.FileHandler("bot.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Загрузка переменных окружения
load_dotenv()

# Токены
TELEGRAM_BOT_TOKEN = "8116572683:AAGxf2ttP-58uts18pRjTIy9cHX0LZfyUsU"
YANDEX_DISK_TOKEN = "y0__wgBELawpyAYkbY0IKiM5oQSstnwahr424ZzdNX_Y9dCWfPK-ac"
LOCAL_TEMP_DIR = "temp_files"

# Проверка токенов
if not TELEGRAM_BOT_TOKEN:
    logger.error("TELEGRAM_BOT_TOKEN не найден. Убедитесь, что он указан в .env файле.")
    sys.exit(1)

if not YANDEX_DISK_TOKEN:
    logger.error("YANDEX_DISK_TOKEN не найден. Убедитесь, что он указан в .env файле.")
    sys.exit(1)

# Логирование токенов для отладки (скрываем часть токена для безопасности)
logger.debug(f"TELEGRAM_BOT_TOKEN: {TELEGRAM_BOT_TOKEN[:10]}...{TELEGRAM_BOT_TOKEN[-10:]}")
logger.debug(f"YANDEX_DISK_TOKEN: {YANDEX_DISK_TOKEN[:10]}...{YANDEX_DISK_TOKEN[-10:]}")

# Создание временной директории
os.makedirs(LOCAL_TEMP_DIR, exist_ok=True)
logger.info(f"Создана временная директория: {LOCAL_TEMP_DIR}")

# Глобальные константы
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DRIVERS_DB_PATH = os.path.join(BASE_DIR, "drivers_db.xlsx")
FIRMS_DB_PATH = os.path.join(BASE_DIR, "firms_db.xlsx")
CARRIERS_DB_PATH = os.path.join(BASE_DIR, "carriers_db.xlsx")
YANDEX_DRIVERS_PATH = "/TransportData/Database/drivers_db.xlsx"
YANDEX_FIRMS_PATH = "/TransportData/Database/firms_db.xlsx"
YANDEX_CARRIERS_PATH = "/TransportData/Database/carriers_db.xlsx"

# Словари для нормализации
CAR_BRANDS = {
    "volvo": "Volvo",
    "scania": "Scania",
    "man": "MAN",
    "daf": "DAF",
    "mercedes": "Mercedes-Benz",
    "mercedes-benz": "Mercedes-Benz",
    "iveco": "Iveco",
    "renault": "Renault",
    "kamaz": "Kamaz",
    "maz": "MAZ",
    "freightliner": "Freightliner",
    "kenworth": "Kenworth",
    "peterbilt": "Peterbilt",
    "isuzu": "Isuzu",
    "hino": "Hino",
    "mitsubishi": "Mitsubishi",
    "fuso": "Fuso",
    "tatra": "Tatra",
    "uaz": "UAZ",
    "gaz": "GAZ",
    "zil": "ZIL",
}

valid_letters = set('АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ')

TRAILER_BRANDS = {
    "schmitz": "Schmitz",
    "krone": "Krone",
    "kögel": "Kögel",
    "kogel": "Kögel",
    "schwarzmüller": "Schwarzmüller",
    "schwarzmuller": "Schwarzmüller",
    "wielton": "Wielton",
    "tonar": "Tonar",
    "grunwald": "Grunwald",
    "kässbohrer": "Kässbohrer",
    "kassbohrer": "Kässbohrer",
    "lamberet": "Lamberet",
    "trailer": "Trailer",
    "nefaz": "Nefaz",
}

COMPOSITE_CITIES = {
    "спб": "Санкт-Петербург", "санкт петербург": "Санкт-Петербург",
    "санкт-петербург": "Санкт-Петербург", "мск": "Москва",
    "нижний новгород": "Нижний Новгород", "усть-лабинск": "Усть-Лабинск",
    "каменск-уральский": "Каменск-Уральский",
    "каменск-шахтинский": "Каменск-Шахтинский",
    "новый уренгой": "Новый Уренгой", "старый оскол": "Старый Оскол",
    "великий новгород": "Великий Новгород",
    "кабардино-балкарской": "Кабардино-Балкарской",
    "прохладненском": "Прохладненском", "г.борисоглебск": "г. Борисоглебск",
    "г.санкт-петербург": "г. Санкт-Петербург",
    "приморско – ахтарского": "Примorsko-Ахтарского",
    "приморско-ахтарского": "Приморско-Ахтарского"
}

CITY_NOMINATIVE = {
    "петрозаводске": "Петрозаводск",
    "москве": "Москва",
    "петербурге": "Санкт-Петербург",
    "борисоглебске": "Борисоглебск",
    "коломне": "Коломна",
}

SMALL_WORDS = {
    "по", "в", "на", "г.", "д.", "кор.", "лит.", "кв.", "р-он", "код", "под.",
    "и", "для", "с", "у", "к", "от", "до"
}

# Инициализация YaDisk с увеличенным таймаутом
y_disk = yadisk.AsyncYaDisk(token=YANDEX_DISK_TOKEN)
logger.info("YaDisk инициализирован")

# Инициализация бота и диспетчера с увеличенным таймаутом
session = AiohttpSession(timeout=120)  # Увеличен таймаут до 120 секунд
bot = Bot(token=TELEGRAM_BOT_TOKEN, session=session)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
logger.info("Telegram Bot и Dispatcher инициализированы")

# Часть 1 завершена

# Часть 2: Блок парсинга (классы состояний и функции парсинга до parse_car_data)

# Классы состояний
class Form(StatesGroup):
    add_driver = State()
    text_input = State()
    confirm_driver = State()
    add_client = State()
    confirm_client = State()
    add_carrier = State()
    confirm_carrier = State()
    add_transportation = State()
    confirm_transportation = State()
    view_clients_search = State()
    view_carriers_search = State()

class DriverStates(StatesGroup):
    waiting_for_add_confirmation = State()
    waiting_for_update_confirmation = State()
    waiting_for_carrier_update_confirmation = State()  # Новое состояние для перевозчиков

# Функция для проверки корректности даты
def validate_date(date_str):
    """
    Проверяет корректность формата даты (дд.мм.гггг).

    Args:
        date_str (str): Строка с датой в формате дд.мм.гггг.

    Returns:
        bool: True, если дата корректна, False в противном случае.
    """
    try:
        day, month, year = map(int, date_str.split('.'))
        if 1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100:
            return True
        logger.warning(f"Некорректная дата: {date_str} (день, месяц или год вне диапазона)")
        return False
    except (ValueError, IndexError) as e:
        logger.warning(f"Ошибка формата даты: {date_str}, ошибка: {str(e)}")
        return False

def parse_passport_issuing_authority(text):
    """
    Извлекает место выдачи паспорта из текста.

    Args:
        text (str): Текст для анализа.

    Returns:
        str or None: Место выдачи паспорта, если найдено, иначе None.
    """
    logger.debug(f"Поиск места выдачи паспорта в тексте: {text[:100]}...")
    passport_place_pattern = re.compile(
        r"(?:паспорт|пасп|п/п|серия\s*и\s*номер|серия).+?(?:выдан|выдано|отделом|кем\s*выдан)\s*"
        r"(?::\s*)?(?:\d{1,2}\.\d{1,2}\.\d{4}(?:г\.?)?\s*)?(.+?)(?=\s*(?:д\.в\.?|дата\s*выдачи|"
        r"код|в/у|ву|водительское\s*удостоверение|права|тел\.?|телефон|а/м|прицеп|полуприцеп|"
        r"п/п|п/пр\.|перевозчик|$))",
        re.IGNORECASE
    )
    passport_place_match = passport_place_pattern.search(text)
    if passport_place_match:
        place = passport_place_match.group(1).strip()
        date_match = re.search(r"(\d{1,2}\.\d{1,2}\.\d{4}(?:г\.?)?)", place)
        if date_match:
            place = place[:date_match.start()].strip()
        code_match = re.search(r"(\d{3}-\d{3})", place)
        if code_match:
            place = place.replace(code_match.group(1), "").strip()
        place = re.sub(
            r"^(выдан|выдано|отделом|кем\s*выдан)\s*",
            "",
            place,
            flags=re.IGNORECASE
        ).strip()
        logger.debug(f"Место выдачи найдено: {place}")
        if len(place) < 5 or place in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            logger.debug("Место выдачи слишком короткое или некорректное, пропускаем")
            return None
        return place

    fallback_pattern = re.compile(
        r"(?:отделом)\s+(.+?)(?=\s*(?:в/у|ву|водительское\s*удостоверение|права|тел\.?|"
        r"телефон|а/м|прицеп|полуприцеп|п/п|п/пр\.|перевозчик|$))",
        re.IGNORECASE
    )
    fallback_match = fallback_pattern.search(text)
    if fallback_match:
        place = fallback_match.group(1).strip()
        logger.debug(f"Место выдачи найдено (fallback): {place}")
        if len(place) < 5 or place in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            logger.debug("Место выдачи слишком короткое или некорректное, пропускаем (fallback)")
            return None
        return place

    logger.debug("Место выдачи паспорта не найдено")
    return None

def parse_phone_numbers(text):
    """
    Извлекает и форматирует номера телефонов из текста.

    Args:
        text (str): Текст для анализа.

    Returns:
        str or None: Строка с номерами телефонов, разделёнными запятыми, или None.
    """
    logger.debug(f"Поиск телефона в тексте: {text}")
    # Очищаем текст от нестандартных пробелов
    text = re.sub(r'\s+', ' ', text).strip()
    phones = []
    
    # Ищем номер ВУ заранее, чтобы использовать его для фильтрации
    vu_match = re.search(
        r"(?:в/у|ву|водительское\s*удостоверение|права|вод\.уд\.)\s*(?:№\s*)?"
        r"([А-ЯЁA-Z]{2}\s*\d{2}\s*\d{6,8})",  # Уточнённый паттерн: 2 буквы + 2 цифры + 6-8 цифр
        text,
        re.IGNORECASE
    )
    vu_number = None
    if vu_match:
        vu_number = re.sub(r"\s+", "", vu_match.group(1))
        logger.debug(f"Найден номер ВУ для фильтрации: {vu_number}")

    # Первый паттерн: ищем номера с явным указанием "тел.", "телефон", "+7" или "8"
    phone_pattern = re.compile(
        r"(?:тел\.?|телефон|\+7|8)\s*[:\-\s]*(\+?\d[\d\s\-\(\)]{9,14})",
        re.IGNORECASE
    )
    phone_matches = phone_pattern.finditer(text)
    for phone_match in phone_matches:
        logger.debug(f"Найден телефон (перед фильтрацией): {phone_match.group(1)}")
        digits = re.sub(r"[^\d]", "", phone_match.group(1))
        # Проверяем длину и формат: телефон должен начинаться с 7 или 8 и иметь 11 цифр
        if len(digits) == 11 and digits[0] in "78":
            phone = f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
            phones.append(phone)
            logger.debug(f"Добавлен телефон: {phone}")
        elif len(digits) == 10:
            phone = f"+7 ({digits[0:3]}) {digits[3:6]}-{digits[6:8]}-{digits[8:10]}"
            phones.append(phone)
            logger.debug(f"Добавлен телефон: {phone}")
        else:
            logger.debug(f"Некорректная длина номера телефона: {digits}")

    # Второй паттерн: ищем номера без явного указания "тел."
    phone_pattern_extra = re.compile(
        r"(?<!\d)(\+?7|8)\s*[\(\-]?\d{3}[\)\-]?\s*\d{3}[\-\s]?\d{2}[\-\s]?\d{2}(?!\d)",  # Уточнённый паттерн для номеров
        re.IGNORECASE
    )
    for phone_match in phone_pattern_extra.finditer(text):
        logger.debug(f"Найден телефон (второй паттерн, перед фильтрацией): {phone_match.group(0)}")
        digits = re.sub(r"[^\d]", "", phone_match.group(0))
        # Проверяем совпадение с номером ВУ
        if vu_number and digits[1:] == vu_number:
            logger.debug(f"Телефон {phone_match.group(0)} совпадает с номером ВУ: {vu_number}")
            continue
        # Исключаем серии паспортов (например, "20 09 118118")
        passport_series_match = re.search(
            r"(?:паспорт|пасп|п/п|серия\s*и\s*номер|серия)\s*(?:серия\s*)?[:\-\s]*(?:№\s*|номер\s*)?(\d{2}\s*\d{2}\s*\d{6})",
            text,
            re.IGNORECASE
        )
        if passport_series_match:
            passport_digits = re.sub(r"[^\d]", "", passport_series_match.group(1))
            if passport_digits == digits:
                logger.debug(f"Телефон {phone_match.group(0)} совпадает с серией паспорта: {passport_digits}")
                continue
        if len(digits) == 11 and digits[0] in "78":
            phone = f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
            if phone not in phones:
                phones.append(phone)
                logger.debug(f"Добавлен телефон (второй паттерн): {phone}")
        elif len(digits) == 10:
            phone = f"+7 ({digits[0:3]}) {digits[3:6]}-{digits[6:8]}-{digits[8:10]}"
            if phone not in phones:
                phones.append(phone)
                logger.debug(f"Добавлен телефон (второй паттерн): {phone}")

    if phones:
        logger.debug(f"Найдены телефоны: {', '.join(phones)}")
        return ', '.join(phones)
    logger.debug("Телефоны не найдены")
    return None

def parse_residence(text):
    """
    Извлекает место жительства или регистрации из текста.

    Args:
        text (str): Текст для анализа.

    Returns:
        str or None: Место жительства, если найдено, иначе None.
    """
    logger.debug(f"Поиск места жительства в тексте: {text[:100]}...")
    residence_match = re.search(
        r'(?:зарегистрирован|проживает|адрес|место жительства|регистрация|прописка)\s*'
        r'[:\-\s]*(.+?)(?=\s*(?:телефон|вод\.|в/у|ву|машина|прицеп|перевозчик|$))',
        text,
        re.IGNORECASE | re.DOTALL
    )
    if residence_match:
        residence = residence_match.group(1).strip()
        logger.debug(f"Место жительства найдено: {residence}")
        return residence

    address_pattern = re.search(
        r'(?:(?:г\.|ул\.|д\.|кор\.|кв\.|р-он)\s*[А-Яа-яЁё0-9\s,-]+?)(?=\s*(?:водительское|'
        r'тел\.?|телефон|в/у|ву|машина|прицеп|перевозчик|$))',
        text,
        re.IGNORECASE
    )
    if address_pattern:
        residence = address_pattern.group(0).strip()
        residence = re.sub(
            r'\b(Води|водительское|удостоверение)\b.*$',
            '',
            residence,
            flags=re.IGNORECASE
        ).strip()
        logger.debug(f"Место жительства найдено (по адресу): {residence}")
        return residence

    logger.debug("Место жительства не найдено")
    return None

def parse_trailer_data(text):
    """
    Извлекает данные о прицепе (бренд и номер).

    Args:
        text (str): Текст для анализа.

    Returns:
        str or None: Данные прицепа (бренд и номер), если найдены, иначе None.
    """
    logger.debug(f"Поиск данных прицепа в тексте: {text[:100]}...")
    # Парсим текст построчно, чтобы сохранить контекст
    lines = text.strip().split('\n')
    for line in lines:
        line = line.strip()
        # Основной паттерн: ищем прицеп с номером, поддерживая пробелы в номере
        trailer_match = re.search(
            r'(?:прицеп|полуприцеп|п/п|п/пр\.)\s*[:\-\s]*(?:([A-Za-zА-Яа-яЁё-]+)\s+)?'
            r'([А-ЯЁ]{2})\s*(\d{4}\s*\d{2}|\d{4,6})',
            line,
            re.IGNORECASE
        )
        if trailer_match:
            brand, letters, number = trailer_match.groups()
            # Очищаем бренд от ключевых слов
            if brand:
                brand = re.sub(r'(прицеп|полуприцеп|п/п|п/пр|рицеп)', '', brand, flags=re.IGNORECASE).strip()
            # Если бренд пустой или является частью ключевого слова, игнорируем его
            if not brand or brand.lower() in ['прицеп', 'полуприцеп', 'п/п', 'п/пр', 'рицеп']:
                brand = ''
            else:
                brand_key = brand.lower() if brand else ''
                brand = TRAILER_BRANDS.get(brand_key, brand.title() if brand else '')
            # Удаляем пробелы из номера для единообразия
            number = number.replace(' ', '')
            # Форматируем номер: первые 2 буквы, затем 4 цифры, затем регион (2 цифры)
            formatted_number = (f"{letters.upper()} {number[:4]} {number[4:]}"
                               if len(number) > 4 else f"{letters.upper()} {number}")
            result = f"{brand} {formatted_number}".strip() if brand else formatted_number
            logger.debug(f"Данные прицепа найдены: {result}")
            return result

    # Дополнительный паттерн: ищем в конце текста
    last_line = lines[-1] if lines else ''
    trailer_match = re.search(
        r'([A-Za-zА-Яа-яЁё-]+)\s+([А-ЯЁA-Z]{2}\s*\d{4}(?:\s*\d{2})?)',
        last_line,
        re.IGNORECASE
    )
    if trailer_match:
        brand, number = trailer_match.groups()
        # Очищаем бренд от ключевых слов
        if brand:
            brand = re.sub(r'(прицеп|полуприцеп|п/п|п/пр|рицеп)', '', brand, flags=re.IGNORECASE).strip()
        # Игнорируем бренд, если он совпадает с ключевыми словами
        if not brand or brand.lower() in ['прицеп', 'полуприцеп', 'п/п', 'п/пр', 'рицеп']:
            brand = ''
        else:
            brand_key = brand.lower() if brand else ''
            brand = TRAILER_BRANDS.get(brand_key, brand.title() if brand else '')
        number = number.replace(' ', '')
        formatted_number = (f"{number[:2].upper()} {number[2:6]} {number[6:]}"
                           if len(number) > 6 else number.upper())
        result = f"{brand} {formatted_number}".strip() if brand else formatted_number
        logger.debug(f"Данные прицепа найдены в последней строке: {result}")
        return result

    logger.debug("Данные прицепа не найдены")
    return None

def parse_car_data(text):
    """
    Извлекает данные об автомобиле (бренд и номер).

    Args:
        text (str): Текст для анализа.

    Returns:
        str or None: Данные автомобиля (бренд и номер), если найдены, иначе None.
    """
    logger.debug(f"Поиск данных автомобиля в тексте: {text[:100]}...")
    car_match = re.search(
        r'(?:машина|авто|автомобиль|а/м|тягач|тс|марка\s*,\s*гос\.?номer)\s*[:\-\s\/]*'
        r'(.+?)(?=\s*(?:прицеп|полуприцеп|п/п|п/пр\.|перевозчик|тел\.?|телефон|$))',
        text,
        re.IGNORECASE
    )
    if car_match:
        car_data = (car_match.group(1).strip()
                    .replace('гос.№', '')
                    .replace('№', '')
                    .replace(';', '')
                    .replace(',', '')
                    .replace('/', ' ')
                    .replace('–', '')  # Удаляем тире
                    .replace('(', '')  # Удаляем скобки
                    .replace(')', '')
                    .replace(':', ''))  # Удаляем двоеточие
        car_data = re.sub(
            r'\b(автомобиль|машина|авто|а/м|мобиль|тягач|марка|гос\.?номер)\b',
            '',
            car_data,
            flags=re.IGNORECASE
        ).strip()
        number_match = re.search(
            r"([А-ЯЁ])\s*(\d{3})\s*([А-ЯЁ]{2})\s*(\d{2,3})$",
            car_data,
            re.IGNORECASE
        )
        if number_match:
            letter1, digits, letters2, region = number_match.groups()
            if all(l.upper() in valid_letters for l in (letter1 + letters2)):
                number = f"{letter1.upper()} {digits} {letters2.upper()} {region}"
                brand = car_data[:number_match.start()].strip()
                logger.debug(f"Данные автомобиля найдены: {brand} {number}")
                return f"{brand} {number}"
        number_match = re.search(
            r"([А-ЯЁ])\s*(\d{3})\s*([А-ЯЁ]{2})(\d{2,3})$",
            car_data,
            re.IGNORECASE
        )
        if number_match:
            letter1, digits, letters2, region = number_match.groups()
            if all(l.upper() in valid_letters for l in (letter1 + letters2)):
                number = f"{letter1.upper()} {digits} {letters2.upper()} {region}"
                brand = car_data[:number_match.start()].strip()
                logger.debug(f"Данные автомобиля найдены: {brand} {number}")
                return f"{brand} {number}"

    last_line = text.strip().split('\n')[-1]
    car_match = re.search(
        r'([A-Za-zА-Яа-яЁё-]+)\s+([А-ЯЁA-Z]\s*\d{3}\s*[А-ЯЁA-Z]{2}\s*\d{2,3})',
        last_line,
        re.IGNORECASE
    )
    if car_match:
        brand, number = car_match.groups()
        brand = re.sub(
            r'\b(автомобиль|машина|авто|а/м|мобиль|тягач|марка|гос\.?номер)\b',
            '',
            brand,
            flags=re.IGNORECASE
        ).strip()
        logger.debug(f"Данные автомобиля найдены в последней строке: {brand} {number}")
        return f"{brand} {number}"

    logger.debug("Данные автомобиля не найдены")
    return None

# Часть 2 завершена

# Часть 3: Блок парсинга (оставшиеся функции)

def normalize_data(data):
    """
    Нормализует данные, приводя их к единому формату.

    Args:
        data (dict): Словарь с данными для нормализации.

    Returns:
        dict: Нормализованные данные.
    """
    logger.debug("Запуск нормализации данных (версия 4.1)")
    normalized_data = {}
    abbreviations = {
        "мвд": "МВД", "уфмс": "УФМС", "оуфмс": "ОУФМС", "ооо": "ООО", "ип": "ИП",
        "ровд": "РОВД", "р-н": "р-н", "г.": "г.", "д.": "д.", "кор.": "кор.",
        "лит.": "лит.", "кв.": "кв.", "снт": "СНТ", "вк": "ВК", "мо": "МО",
        "рк": "РК", "ул.": "ул.", "обл.": "обл.", "респ.": "респ.", "пгт": "пгт",
        "с.": "с.", "нп": "нп", "мкр.": "мкр.", "тер.": "тер.", "мро": "МРО",
        "тп": "ТП", "овд": "ОВД", "оао": "ОАО", "зао": "ЗАО"
    }

    for key, value in data.items():
        if (value is None or
                (isinstance(value, str) and
                 (not value.strip() or value.lower() in ["не указано", "не указан"]))):
            continue  # Пропускаем поля, которые не заполнены

        if key in ["Цена", "Оплата"]:
            try:
                normalized_data[key] = float(value)
            except (ValueError, TypeError):
                logger.warning(f"Некорректное значение для {key}: {value}")
                normalized_data[key] = 0
            continue

        preserve_case_fields = [
            "Паспорт_дата_выдачи", "Паспорт_код_подразделения",
            "ВУ_серия_и_номер", "Дата_рождения", "Дата_перевозки",
            "В/У_дата_срок", "ИНН", "Контакт"
        ]
        if key in preserve_case_fields:
            normalized_data[key] = value
            continue

        if not isinstance(value, str):
            value = str(value)
        value_lower = value.lower()

        if key in [
            "Водитель", "Паспорт_место_выдачи", "Адрес_регистрации",
            "Место_рождения", "Место_жительства", "Гражданство",
            "Фирма", "Направление", "Перевозчик", "Клиент",
            "Краткое название", "Полное название", "Имя перевозчика"
        ]:
            logger.debug(f"Before normalization ({key}): {value}")
            words = value.split()
            corrected_words = []
            i = 0
            while i < len(words):
                word = words[i]
                word_lower = word.lower()
                logger.debug(f"Processing word: {word}, lower: {word_lower}")

                if word_lower == "д." and i > 0:
                    prev_word = words[i - 1].lower()
                    if prev_word == "ул.":
                        corrected_words.append("д.")
                    else:
                        corrected_words.append("д.")
                    i += 1
                    continue

                composite_key = (" ".join(words[i:i+2]).lower()
                                if i + 1 < len(words) else word_lower)
                if composite_key in COMPOSITE_CITIES:
                    corrected_words.append(COMPOSITE_CITIES[composite_key])
                    i += 2
                    continue

                word_clean = re.sub(r'[^\w\s-]', '', word).lower()
                if word_clean in abbreviations:
                    target = abbreviations[word_clean]
                    word_without_punctuation = re.sub(r'[^\w\s-]', '', word)
                    if word_without_punctuation == target:
                        corrected_words.append(word)
                    else:
                        punctuation = re.sub(r'[\w\s-]', '', word)
                        corrected_words.append(target + punctuation)
                    i += 1
                    continue

                if word_lower in SMALL_WORDS:
                    corrected_words.append(word_lower)
                    i += 1
                    continue

                if word_lower in ["районе", "области"]:
                    corrected_words.append(word_lower)
                    i += 1
                    continue

                if (word_lower.startswith("г.") or
                        word_lower.startswith("д.") or
                        word_lower.startswith("кв.")):
                    corrected_words.append(word)
                    i += 1
                    continue

                if '-' in word:
                    corrected_words.append(word.title())
                    i += 1
                    continue

                if key == "Место_жительства" and word_lower in CITY_NOMINATIVE:
                    corrected_words.append(CITY_NOMINATIVE[word_lower])
                    i += 1
                    continue

                if key in ["Перевозчик", "Клиент", "Полное название", "Краткое название", "Имя перевозчика"]:
                    if word_lower == "ип" and i + 1 < len(words):
                        corrected_words.append("ИП")
                        corrected_words.append(words[i + 1].title())
                        i += 2
                        continue
                    elif word_lower == "ооо" and i + 1 < len(words):
                        corrected_words.append("ООО")
                        corrected_words.append(words[i + 1].title())
                        i += 2
                        continue
                    elif word_lower == "оао" and i + 1 < len(words):
                        corrected_words.append("ОАО")
                        corrected_words.append(words[i + 1].title())
                        i += 2
                        continue
                    elif word_lower == "зао" and i + 1 < len(words):
                        corrected_words.append("ЗАО")
                        corrected_words.append(words[i + 1].title())
                        i += 2
                        continue

                corrected_words.append(word.title())
                i += 1

            normalized_value = " ".join(corrected_words)

            if key == "Паспорт_место_выдачи":
                stop_words = [
                    "а/м", "прицеп", "полуприцеп", "п/п", "п/пр.",
                    "перевозчик", "телефон", "тел."
                ]
                for stop_word in stop_words:
                    stop_index = normalized_value.lower().find(stop_word)
                    if stop_index != -1:
                        normalized_value = normalized_value[:stop_index].strip()
                        break

            logger.debug(f"After normalization ({key}): {normalized_value}")
            normalized_data[key] = normalized_value

        elif key == "Телефон":
            normalized_data[key] = value

        elif key == "Марка_машины_номер":
            logger.debug(f"Before normalization (Марка_машины_номер): {value}")
            parts = value.split()
            if len(parts) >= 2:
                number_parts = []
                found_number = False
                for i in range(len(parts) - 1, -1, -1):
                    part = parts[i]
                    if not found_number:
                        if re.match(r"^\d{2,3}$", part):
                            number_parts.insert(0, part)
                            found_number = True
                    else:
                        if (len(number_parts) == 1 and
                                re.match(r"^[А-ЯЁA-Z]{2,3}$", part) and
                                not any(c.isdigit() for c in part)):
                            number_parts.insert(0, part)
                        elif len(number_parts) == 2 and re.match(r"^\d{3}$", part):
                            number_parts.insert(0, part)
                        elif (len(number_parts) == 3 and
                              re.match(r"^[А-ЯЁA-Z]$", part) and
                              not any(c.isdigit() for c in part)):
                            number_parts.insert(0, part)
                        else:
                            break
                if number_parts:
                    number = ' '.join(number_parts).upper()
                    logger.debug(f"Extracted number: {number}")
                    brand_parts = (parts[:-len(number_parts)]
                                   if parts[:-len(number_parts)] else [parts[0]])
                    brand = ' '.join(brand_parts).lower()
                    brand = re.sub(
                        r'\b(мобиль|автомобиль|авто|а/м|машина|тягач|марка|гос\.?номер)\b',
                        '',
                        brand,
                        flags=re.IGNORECASE
                    ).strip()
                    brand = re.sub(r'[^a-zA-Zа-яА-ЯёЁ\s]', '', brand).strip()
                    brand_key = re.sub(r'[^a-zA-Zа-яА-ЯёЁ]', '', brand.lower())
                    logger.debug(f"Brand: {brand}, Brand Key: {brand_key}")
                    normalized_brand = CAR_BRANDS.get(brand_key, brand.title())
                    logger.debug(f"Normalized brand from CAR_BRANDS: {normalized_brand}")
                    normalized_value = f"{normalized_brand} {number}".strip()
                    logger.debug(f"After normalization (Марка_машины_номер): {normalized_value}")
                    normalized_data[key] = normalized_value
                else:
                    normalized_data[key] = value.upper()
            else:
                normalized_data[key] = value.upper()

        elif key == "Прицеп_номер":
            logger.debug(f"Before normalization (Прицеп_номер): {value}")
            parts = value.split()
            # Проверяем, является ли первая часть брендом
            first_part = parts[0].lower() if parts else ''
            if len(parts) > 2 and first_part in TRAILER_BRANDS:  # Если есть бренд (например, "Schmitz ЕТ 1913 50")
                brand = TRAILER_BRANDS[first_part]
                number = ' '.join(parts[1:]).upper()
                number_match = re.match(
                    r"([А-ЯЁ]{2})\s*(\d{4,6}(?:\s*\d{2})?)$",
                    number
                )
                if number_match:
                    letters, digits = number_match.groups()
                    normalized_number = f"{letters} {digits}".strip()
                    normalized_data[key] = f"{brand} {normalized_number}".strip()
                else:
                    logger.warning(f"Некорректный формат номера прицепа: {number}")
                    normalized_data[key] = f"{brand} {number}".strip()
            else:  # Если это только номер (например, "ЕТ 1913 50")
                normalized_data[key] = value.upper()  # Сохраняем регистр букв как есть
            logger.debug(f"After normalization (Прицеп_номер): {normalized_data[key]}")

        else:
            normalized_data[key] = value

    return normalized_data

def parse_by_keywords(text, is_driver_data=False):
    """
    Парсит текст по ключевым словам и извлекает данные.

    Args:
        text (str): Текст для парсинга.
        is_driver_data (bool): Флаг, указывающий, являются ли данные данными водителя.

    Returns:
        tuple: (raw_data, normalized_data) - сырые и нормализованные данные.
    """
    data = {}
    text = text.strip()

    if not text:
        logger.error("Получен пустой текст для парсинга")
        return data, data

    text = re.sub(r"\s*\n\s*", "\n", text).strip()
    full_text = text.replace('\n', ' ')
    logger.debug(f"Полный текст для парсинга: {full_text}")

    abbreviations = {
        "мвд": "МВД", "уфмс": "УФМС", "ооо": "ООО", "ип": "ИП", "ровд": "РОВД",
        "р-н": "р-н", "г.": "г.", "д.": "д.", "кор.": "кор.", "лит.": "лит.",
        "кв.": "кв.", "снт": "СНТ", "вк": "ВК", "ул.": "ул.", "обл.": "обл.",
        "респ.": "респ.", "пгт": "пгт", "с.": "с.", "нп": "нп", "мкр.": "мкр.",
        "тер.": "тер.", "мро": "МРО", "тп": "ТП", "овд": "ОВД", "оао": "ОАО", "зао": "ЗАО"
    }

    def parse_date(pattern, field_name, compare_field=None):
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            date_str = match.group(1)
            if validate_date(date_str):
                date_parts = date_str.split('.')
                formatted_date = f"{date_parts[0].zfill(2)}.{date_parts[1].zfill(2)}.{date_parts[2]}"
                if compare_field and compare_field in data and formatted_date == data[compare_field]:
                    logger.warning(f"{field_name} совпадает с {compare_field}: {formatted_date}, пропускаем")
                    return None
                return formatted_date
            logger.warning(f"Некорректный формат {field_name}: {date_str}")
        return None

    # Парсинг ФИО водителя
    # Сначала пробуем строгий формат с ключевыми словами "водитель", "ф.и.о. водителя" или "фио"
    fio_match = re.search(
        r"(?:водитель|ф\.и\.о\. водителя|фио)\s*[:\-\s]*([А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+){1,3})[\.\s]*(?=\s*(?:паспорт|данные|тел\.?|телефон|машина|в/у|ву|вод\.уд\.|права|$))",
        full_text,
        re.IGNORECASE
    )
    if fio_match:
        fio = re.sub(r'\s+', ' ', fio_match.group(1)).strip()
        carrier_match = re.search(
            r"(?:перевозчик|превозчик)\s*[:\-\s]*(.+?)(?=\s*(?:регистрация|адрес|тел\.?|водитель|паспорт|машина|$))",
            full_text,
            re.IGNORECASE
        )
        carrier_fio = None
        if carrier_match:
            carrier_text = carrier_match.group(1).strip()
            carrier_fio_match = re.search(
                r"([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)|(?:ип\s+([А-ЯЁ][а-яё]+))",
                carrier_text,
                re.IGNORECASE
            )
            if carrier_fio_match:
                if carrier_fio_match.group(1):
                    carrier_fio = re.sub(r'\s+', ' ', carrier_fio_match.group(1)).strip()
                elif carrier_fio_match.group(2):
                    carrier_fio = f"ИП {carrier_fio_match.group(2)}"

        if carrier_fio and fio.lower() == carrier_fio.lower():
            logger.debug(f"ФИО {fio} совпадает с ФИО перевозчика, пропускаем")
        elif not any(keyword.lower() in fio.lower() for keyword in [
            "данные", "машина", "паспорт", "ровд", "уфмс", "отделом", "выдан",
            "области", "республике", "перевозчик", "ип"
        ]):
            data["Водитель"] = fio
            logger.debug(f"ФИО найдено: {data['Водитель']}")
        else:
            logger.debug(f"ФИО содержит ключевые слова, пропускаем: {fio}")
    else:
        # Если строгий формат не сработал, пробуем угадать ФИО
        # Сначала проверяем перевозчика (чтобы исключить совпадение)
        carrier_match = re.search(
            r"(?:перевозчик|превозчик)\s*[:\-\s]*(.+?)(?=\s*(?:регистрация|адрес|тел\.?|водитель|паспорт|машина|$))",
            full_text,
            re.IGNORECASE
        )
        carrier_fio = None
        if carrier_match:
            carrier_text = carrier_match.group(1).strip()
            carrier_fio_match = re.search(
                r"([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)|(?:ип\s+([А-ЯЁ][а-яё]+))",
                carrier_text,
                re.IGNORECASE
            )
            if carrier_fio_match:
                if carrier_fio_match.group(1):
                    carrier_fio = re.sub(r'\s+', ' ', carrier_fio_match.group(1)).strip()
                elif carrier_fio_match.group(2):
                    carrier_fio = f"ИП {carrier_fio_match.group(2)}"

        # Пробуем угадать ФИО: либо в начале текста, либо после любых символов
        fio_guess = None
        # Проверяем начало текста, но исключаем случаи с "ООО", "ИП" и т.д.
        if not re.match(r"^(?:ООО|ИП|ОАО|ЗАО)\s+", full_text, re.IGNORECASE):
            fio_guess = re.search(
                r"^(?:[^а-яА-ЯёЁ]*)([А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+){1,2})[\.\s]*(?=\s*(?:паспорт|данные|тел\.?|телефон|машина|в/у|ву|вод\.уд\.|права|$))",
                full_text,
                re.IGNORECASE
            )
        # Если начало текста не подходит, пробуем старый запасной вариант
        if not fio_guess:
            fio_guess = re.search(
                r"([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+){1,2})[\.\s]*(?=\s*(?:паспорт|данные|тел\.?|телефон|машина|в/у|ву|вод\.уд\.|права|$))",
                full_text,
                re.IGNORECASE
            )

        if fio_guess:
            fio = re.sub(r'\s+', ' ', fio_guess.group(1)).strip()
            if carrier_fio and fio.lower() == carrier_fio.lower():
                logger.debug(f"Угаданное ФИО {fio} совпадает с ФИО перевозчика, пропускаем")
            elif not any(keyword.lower() in fio.lower() for keyword in [
                "данные", "машина", "паспорт", "ровд", "уфмс", "отделом", "выдан",
                "области", "республике", "перевозчик", "ип"
            ]):
                data["Водитель"] = fio
                logger.debug(f"ФИО угадано: {data['Водитель']}")
            else:
                logger.debug(f"Угаданное ФИО содержит ключевые слова: {fio}")
        else:
            logger.warning(f"ФИО водителя не найдено в тексте: {full_text[:50]}...")

    pass_series_number_pattern = re.compile(
        r"(?:паспорт|пасп|п/п|серия\s*и\s*номер|серия|данные\s*водителя)\s*(?:серия\s*)?"
        r"[:\-\s]*(?:№\s*|номер\s*)?(\d{2}\s*\d{2}|\d{4})\s*(?:№\s*|номер\s*)?(\d{6})",
        re.IGNORECASE
    )
    passport_series_number_match = pass_series_number_pattern.search(full_text)
    if passport_series_number_match:
        series = passport_series_number_match.group(1).replace(' ', '')
        number = passport_series_number_match.group(2)
        data["Паспорт_серия_и_номер"] = f"{series[:2]} {series[2:]} № {number}"
        logger.debug(f"Паспорт серия и номер: {data['Паспорт_серия_и_номер']}")
    else:
        logger.debug(f"Серия и номер паспорта не найдены в тексте: {full_text[:50]}...")

    passport_place = parse_passport_issuing_authority(full_text)
    if passport_place:
        data["Паспорт_место_выдачи"] = passport_place
        logger.debug(f"Место выдачи паспорта: {data['Паспорт_место_выдачи']}")
    else:
        logger.debug(f"Место выдачи паспорта не найдено в тексте: {full_text[:50]}...")

    if "Паспорт_дата_выдачи" not in data:
        date_near_passport = re.search(
            r"(?:паспорт|пасп|п/п|серия\s*и\s*номер|серия|данные\s*водителя).+?"
            r"(?:выдан|выдано|отделом|кем\s*выдан).+?(\d{1,2}\.\d{1,2}\.\d{4}(?:г\.?)?)",
            full_text,
            re.IGNORECASE
        )
        if date_near_passport:
            date_str = re.sub(r'г\.?$', '', date_near_passport.group(1))
            if validate_date(date_str):
                date_parts = date_str.split('.')
                data["Паспорт_дата_выдачи"] = (f"{date_parts[0].zfill(2)}."
                                               f"{date_parts[1].zfill(2)}.{date_parts[2]}")
                logger.debug(f"Дата выдачи паспорта найдена рядом с 'паспорт': "
                             f"{data['Паспорт_дата_выдачи']}")
        else:
            data["Паспорт_дата_выдачи"] = parse_date(
                r"(?:д\.в\.?|дата\s*выдачи|выдан|выдано)\s*[,:\-\s]*(\d{1,2}\.\d{1,2}\.\d{4}(?:г\.?)?)",
                "Паспорт_дата_выдачи"
            )
    if not data.get("Паспорт_дата_выдачи"):
        logger.debug(f"Дата выдачи паспорта не найдена в тексте: {full_text[:50]}...")

    if "Паспорт_код_подразделения" not in data:
        passport_code_match = re.search(
            r"(?:код\s*(?:подразделения)?\s*[:\-\s]*|\s)(\d{3}-\d{3})",
            full_text,
            re.IGNORECASE
        )
        if passport_code_match:
            data["Паспорт_код_подразделения"] = passport_code_match.group(1)
            logger.debug(f"Код подразделения паспорта: {data['Паспорт_код_подразделения']}")
        else:
            logger.debug(f"Код подразделения паспорта не найден в тексте: {full_text[:50]}...")

    vu_pattern = re.compile(
        r"(?:в/у|ву|водительское\s*удостоверение|права|вод\.уд\.)\s*(?:№\s*)?"
        r"([А-ЯЁA-Z\d\s]{2,4}\s*\d{6,8})(?:\s*(?:выдан|выдано|от)\s*"
        r"(\d{1,2}\.\d{1,2}\.\d{4}(?:г\.?)?))?",
        re.IGNORECASE
    )
    vu_match = vu_pattern.search(full_text)
    if vu_match:
        vu_number = re.sub(r'\s+', '', vu_match.group(1)).upper()
        logger.debug(f"Найден необработанный номер ВУ: {vu_number}")
        if len(vu_number) == 10:
            if re.match(r"[А-ЯЁA-Z]{2}\d{2}\d{6}", vu_number):
                formatted_vu = f"{vu_number[:2]} {vu_number[2:4]} {vu_number[4:]}"
            elif re.match(r"\d{2}\d{2}\d{6}", vu_number):
                formatted_vu = f"{vu_number[:2]} {vu_number[2:4]} {vu_number[4:]}"
            else:
                formatted_vu = vu_number
        elif re.match(r"[А-ЯЁA-Z]{2}\d{6,8}", vu_number):
            formatted_vu = vu_number[:2] + " " + vu_number[2:]
        else:
            formatted_vu = vu_number
        data["ВУ_серия_и_номер"] = formatted_vu
        vu_date = vu_match.group(2)
        if vu_date:
            vu_date = re.sub(r'г\.?$', '', vu_date)
            if validate_date(vu_date):
                date_parts = vu_date.split('.')
                data["В/У_дата_срок"] = (f"выдан {date_parts[0].zfill(2)}."
                                         f"{date_parts[1].zfill(2)}.{date_parts[2]}")
        logger.debug(f"Водительское удостоверение: {data['ВУ_серия_и_номер']}")
    else:
        vu_fallback = re.search(
            r"(?:в/у|ву|водительское\s*удостоверение|права|вод\.уд\.)\s*(?:№\s*)?"
            r"(\d{2}\s*\d{2}\s*\d{6,8})",
            full_text,
            re.IGNORECASE
        )
        if vu_fallback:
            vu_number = re.sub(r'\s+', '', vu_fallback.group(1)).upper()
            formatted_vu = f"{vu_number[:2]} {vu_number[2:4]} {vu_number[4:]}"
            data["ВУ_серия_и_номер"] = formatted_vu
            logger.debug(f"Водительское удостоверение (fallback): {data['ВУ_серия_и_номер']}")
        else:
            logger.debug(f"Водительское удостоверение не найдено в тексте: {full_text[:50]}...")

    data["Дата_рождения"] = parse_date(
        r"(?:д\.р\.?|дата\s*рождения|рождения)\s*[:\-\s\.]*(\d{1,2}\.\d{1,2}\.\d{4})",
        "Дата_рождения",
        "Паспорт_дата_выдачи"
    )
    if not data.get("Дата_рождения"):
        logger.debug(f"Дата рождения не найдена в тексте: {full_text[:50]}...")

    birth_place_match = re.search(
        r"(?:место\s*рождения|рождения)\s*[:\-\s\.]*([А-ЯЁа-яё\s,.-]+?)(?=\s*(?:паспорт|"
        r"тел\.?|машина|$))",
        full_text,
        re.IGNORECASE
    )
    if birth_place_match:
        data["Место_рождения"] = birth_place_match.group(1).strip()
        logger.debug(f"Место рождения: {data['Место_рождения']}")
    else:
        logger.debug(f"Место рождения не найдено в тексте: {full_text[:50]}...")

    citizenship_match = re.search(
        r"(?:гр\.?|гражданство)\s*[:\-\s]*([А-ЯЁа-яё\s-]+?)(?=\s*(?:м\.ж\.?|тел\.?|машина|$))",
        full_text,
        re.IGNORECASE
    )
    if citizenship_match:
        data["Гражданство"] = citizenship_match.group(1).strip()
        logger.debug(f"Гражданство: {data['Гражданство']}")
    else:
        logger.debug(f"Гражданство не найдено в тексте: {full_text[:50]}...")

    residence = parse_residence(full_text)
    if residence:
        data["Место_жительства"] = residence
        logger.debug(f"Место жительства найдено: {data['Место_жительства']}")
    else:
        logger.debug(f"Место жительства не найдено в тексте: {full_text}")

    phones = parse_phone_numbers(full_text)
    if phones:
        data["Телефон"] = phones
        logger.debug(f"Телефон: {data['Телефон']}")
    else:
        logger.debug(f"Телефон не найден в тексте: {full_text[:50]}...")

    car_data = parse_car_data(full_text)
    if car_data:
        data["Марка_машины_номер"] = car_data
        logger.debug(f"Car data: {data['Марка_машины_номер']}")
    else:
        logger.debug(f"Данные об автомобиле не найдены в тексте: {full_text[:50]}...")

    trailer_data = parse_trailer_data(text)  # Передаём оригинальный текст, а не full_text
    if trailer_data:
        data["Прицеп_номер"] = trailer_data
        logger.debug(f"Прицепы: {data['Прицеп_номер']}")
    else:
        logger.debug(f"Прицепы не найдены в тексте: {full_text[:50]}...")

    # Улучшенный парсинг для клиента
    client_match = re.search(
        r"(?:клиент|заказчик)\s*[:\-\s]*(.+?)(?=\s*(?:перевозчик|водитель|фирма|направление|тел\.?|машина|$))",
        full_text,
        re.IGNORECASE
    )
    if client_match:
        data["Клиент"] = client_match.group(1).strip()
        logger.debug(f"Клиент: {data['Клиент']}")
    else:
        logger.debug(f"Клиент не найден в тексте: {full_text[:50]}...")

    # Улучшенный парсинг для перевозчика
    carrier_match = re.search(
        r"(?:перевозчик|превозчик)\s*[:\-\s]*(.+?)(?=\s*(?:регистрация|адрес|тел\.?|"
        r"водитель|паспорт|машина|$))",
        full_text,
        re.IGNORECASE
    )
    if carrier_match:
        carrier_text = carrier_match.group(1).strip()
        data["Перевозчик"] = carrier_text
        logger.debug(f"Перевозчик: {data['Перевозчик']}")

        # Парсинг имени перевозчика
        carrier_name_match = re.search(
            r"(?:ип\s+([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+))|(?:ип\s+([А-ЯЁ][а-яё]+))",
            carrier_text,
            re.IGNORECASE
        )
        if carrier_name_match:
            if carrier_name_match.group(1):
                data["Имя перевозчика"] = carrier_name_match.group(1).strip()
            elif carrier_name_match.group(2):
                data["Имя перевозчика"] = carrier_name_match.group(2).strip()
            logger.debug(f"Имя перевозчика: {data.get('Имя перевозчика', 'Не указано')}")
        else:
            logger.debug("Имя перевозчика не указано или перевозчик не является ИП")

        # Парсинг телефона перевозчика
        # Ищем телефон в части текста после названия перевозчика
        carrier_phone_match = re.search(
            r"(?:перевозчик|превозчик)\s*[:\-\s]*(.+?)(?=\s*(?:водитель|паспорт|машина|$))",
            full_text,
            re.IGNORECASE
        )
        if carrier_phone_match:
            carrier_section = carrier_phone_match.group(1).strip()
            carrier_phone = parse_phone_numbers(carrier_section)
            if carrier_phone:
                data["Контакт"] = carrier_phone
                logger.debug(f"Контакт перевозчика: {data['Контакт']}")
            else:
                logger.debug("Телефон перевозчика не найден в секции перевозчика")
        else:
            logger.debug("Секция перевозчика не найдена для парсинга телефона")

    else:
        logger.debug(f"Перевозчик не найден в тексте: {full_text[:50]}...")

    if not is_driver_data:
        for field, pattern in [
            ("Фирма", r"фирма\s+([\w\s-]+)"),
            ("Направление", r"направление\s+([\w\s-]+)"),
            ("Цена", r"цена\s+(\d+)"),
            ("Оплата", r"оплата\s+(\d+)"),
            ("Дата_перевозки", r"дата\s+перевозки\s+(\d{2}\.\d{2}\.\d{4})"),
            ("Пометка", r"пометка\s+([\w\s-]+)")
        ]:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                data[field] = (int(match.group(1))
                               if field in ["Цена", "Оплата"]
                               else match.group(1).strip())
                logger.debug(f"{field}: {data[field]}")
            else:
                logger.debug(f"{field} не найдена в тексте: {full_text[:50]}...")

    raw_data = data.copy()
    logger.debug(f"Raw data before normalization: {raw_data}")

    data = normalize_data(data)
    logger.debug(f"Распарсенные данные после нормализации: {data}")

    return raw_data, data

def format_driver_data(data):
    """
    Форматирует данные в читаемую строку.

    Args:
        data (dict): Словарь с данными.

    Returns:
        str: Отформатированная строка с данными.
    """
    result = []
    for key, value in data.items():
        if value is None or key.endswith('_id'):  # Пропускаем поля с ID
            continue
        # Убираем лишние слова из ключей
        formatted_key = key.replace('_', ' ')
        # Специальная обработка для "Прицеп номер"
        if key == 'Прицеп_номер':
            formatted_key = 'Прицеп'
        result.append(f"{formatted_key}: {value}")
    formatted_result = "\n".join(result)
    logger.debug(f"Formatted driver data: {formatted_result}")
    return formatted_result

# Часть 3 завершена



# Часть 4: Классы DatabaseManager, YandexDiskUploader, ExcelManager

class DatabaseManager:
    """Класс для управления базами данных водителей, фирм и перевозчиков на Яндекс.Диске."""

    def __init__(self, y_disk, local_dir):
        """
        Инициализирует объект DatabaseManager.

        Args:
            y_disk (yadisk.AsyncYaDisk): Клиент Яндекс.Диска.
            local_dir (str): Локальная директория для хранения файлов.

        Raises:
            ValueError: Если y_disk не является объектом yadisk.AsyncYaDisk.
        """
        if not isinstance(y_disk, yadisk.AsyncYaDisk):
            logger.error("y_disk должен быть объектом yadisk.AsyncYaDisk")
            raise ValueError("y_disk должен быть объектом yadisk.AsyncYaDisk")
        if not os.path.exists(local_dir):
            os.makedirs(local_dir, exist_ok=True)
            logger.info(f"Создана локальная директория: {local_dir}")
        self.y_disk = y_disk
        self.local_dir = local_dir
        self.drivers_db_path = os.path.join(local_dir, 'drivers_db.xlsx')
        self.firms_db_path = os.path.join(local_dir, 'firms_db.xlsx')
        self.carriers_db_path = os.path.join(local_dir, 'carriers_db.xlsx')
        self.remote_drivers_db = '/TransportData/Database/drivers_db.xlsx'
        self.remote_firms_db = '/TransportData/Database/firms_db.xlsx'
        self.remote_carriers_db = '/TransportData/Database/carriers_db.xlsx'
        self.drivers_wb = None
        self.firms_wb = None
        self.carriers_wb = None
        # Добавляем флаги для отслеживания изменений
        self.drivers_modified = False
        self.firms_modified = False
        self.carriers_modified = False
        logger.info("DatabaseManager инициализирован")

    async def ensure_path_exists(self, path):
        """
        Проверяет и создаёт путь на Яндекс.Диске, если он не существует.

        Args:
            path (str): Путь на Яндекс.Диске.

        Raises:
            ValueError: Если path не является строкой.
            Exception: Если не удалось создать директорию.
        """
        if not isinstance(path, str):
            logger.error("path должен быть строкой")
            raise ValueError("path должен быть строкой")
        parts = path.strip('/').split('/')
        current_path = ''
        for part in parts:
            if part:
                current_path += f'/{part}'
                exists = await self.y_disk.exists(current_path)
                if not exists:
                    try:
                        await self.y_disk.mkdir(current_path)
                        logger.info(f"Создана директория на Яндекс.Диске: {current_path}")
                    except Exception as e:
                        logger.error(f"Ошибка создания директории {current_path}: {str(e)}")
                        raise

    async def download_db(self, remote_path, local_path, retries=3, timeout=60):
        """
        Скачивает базу данных с Яндекс.Диска с несколькими попытками.

        Args:
            remote_path (str): Путь к файлу на Яндекс.Диске.
            local_path (str): Локальный путь для сохранения файла.
            retries (int): Количество попыток.
            timeout (int): Таймаут для операции.

        Raises:
            ValueError: Если пути не являются строками.
            Exception: Если не удалось скачать файл.
        """
        if not isinstance(remote_path, str) or not isinstance(local_path, str):
            logger.error("remote_path и local_path должны быть строками")
            raise ValueError("remote_path и local_path должны быть строками")
        for attempt in range(retries):
            try:
                logger.debug(f"Попытка {attempt + 1}/{retries} скачать {remote_path}")
                exists = await self.y_disk.exists(remote_path)
                if exists:
                    await self.y_disk.download(remote_path, local_path, timeout=timeout)
                    logger.info(f"Скачан файл: {remote_path} -> {local_path}")
                    return
                else:
                    logger.warning(f"Файл {remote_path} не найден на Яндекс.Диске")
                    return
            except Exception as e:
                logger.error(f"Ошибка скачивания {remote_path} (попытка {attempt + 1}): {str(e)}")
                if attempt < retries - 1:
                    await asyncio.sleep(5)
                else:
                    logger.error(f"Не удалось скачать {remote_path} после {retries} попыток")
                    raise

    async def load_dbs(self):
        """
        Загружает базы данных в память.

        Raises:
            FileNotFoundError: Если файл базы данных не найден.
            Exception: Если произошла ошибка при загрузке баз данных.
        """
        try:
            if not self.drivers_wb:
                logger.debug(f"Загрузка базы данных водителей: {self.drivers_db_path}")
                self.drivers_wb = load_workbook(self.drivers_db_path)
                logger.debug(f"Загружен файл базы данных водителей: {self.drivers_db_path}, строк: {self.drivers_wb.active.max_row}")
            if not self.firms_wb:
                logger.debug(f"Загрузка базы данных фирм: {self.firms_db_path}")
                self.firms_wb = load_workbook(self.firms_db_path)
                logger.debug(f"Загружен файл базы данных фирм: {self.firms_db_path}, строк: {self.firms_wb.active.max_row}")
            if not self.carriers_wb:
                logger.debug(f"Загрузка базы данных перевозчиков: {self.carriers_db_path}")
                self.carriers_wb = load_workbook(self.carriers_db_path)
                logger.debug(f"Загружен файл базы данных перевозчиков: {self.carriers_db_path}, строк: {self.carriers_wb.active.max_row}")
        except FileNotFoundError as e:
            logger.error(f"Файл базы данных не найден: {e}")
            raise
        except Exception as e:
            logger.error(f"Ошибка при загрузке баз данных: {e}")
            raise

    async def save_dbs(self):
        """
        Сохраняет только изменённые базы данных локально и загружает их на Яндекс.Диск.

        Raises:
            Exception: Если произошла ошибка при сохранении баз данных.
        """
        try:
            if self.drivers_modified and self.drivers_wb:
                logger.debug(f"Сохранение базы данных водителей локально: {self.drivers_db_path}")
                self.drivers_wb.save(self.drivers_db_path)
                logger.debug(f"Загрузка базы данных водителей на Яндекс.Диск: {self.remote_drivers_db}")
                await self.y_disk.upload(
                    self.drivers_db_path,
                    self.remote_drivers_db,
                    overwrite=True
                )
                logger.info(
                    f"Сохранена база данных водителей: {self.drivers_db_path} -> "
                    f"{self.remote_drivers_db}"
                )
                self.drivers_modified = False  # Сбрасываем флаг

            if self.firms_modified and self.firms_wb:
                logger.debug(f"Сохранение базы данных фирм локально: {self.firms_db_path}")
                self.firms_wb.save(self.firms_db_path)
                logger.debug(f"Загрузка базы данных фирм на Яндекс.Диск: {self.remote_firms_db}")
                await self.y_disk.upload(
                    self.firms_db_path,
                    self.remote_firms_db,
                    overwrite=True
                )
                logger.info(
                    f"Сохранена база данных фирм: {self.firms_db_path} -> "
                    f"{self.remote_firms_db}"
                )
                self.firms_modified = False  # Сбрасываем флаг

            if self.carriers_modified and self.carriers_wb:
                logger.debug(f"Сохранение базы данных перевозчиков локально: {self.carriers_db_path}")
                self.carriers_wb.save(self.carriers_db_path)
                logger.debug(f"Загрузка базы данных перевозчиков на Яндекс.Диск: {self.remote_carriers_db}")
                await self.y_disk.upload(
                    self.carriers_db_path,
                    self.remote_carriers_db,
                    overwrite=True
                )
                logger.info(
                    f"Сохранена база данных перевозчиков: {self.carriers_db_path} -> "
                    f"{self.remote_carriers_db}"
                )
                self.carriers_modified = False  # Сбрасываем флаг

        except Exception as e:
            logger.error(f"Ошибка при сохранении баз данных: {str(e)}")
            raise

    async def ensure_dbs_exist(self):
        """
        Проверяет существование баз данных и создаёт их при необходимости.

        Raises:
            Exception: Если не удалось создать или загрузить базы данных.
        """
        logger.debug("Обеспечение существования баз данных...")
        await self.ensure_path_exists('/TransportData/Database')
        await self.download_db(
            self.remote_drivers_db,
            self.drivers_db_path,
            retries=3,
            timeout=60
        )
        await self.download_db(
            self.remote_firms_db,
            self.firms_db_path,
            retries=3,
            timeout=60
        )
        await self.download_db(
            self.remote_carriers_db,
            self.carriers_db_path,
            retries=3,
            timeout=60
        )
        await self.load_dbs()
        logger.debug("Базы данных успешно подготовлены")

    async def lookup_driver(self, driver_id):
        """
        Ищет водителя в базе данных по ID.

        Args:
            driver_id (int): ID водителя для поиска.

        Returns:
            dict: Данные водителя, если найден, иначе None.

        Raises:
            ValueError: Если driver_id не является числом.
            Exception: Если произошла ошибка при поиске.
        """
        if not isinstance(driver_id, int):
            logger.error("driver_id должен быть числом")
            raise ValueError("driver_id должен быть числом")
        try:
            await self.load_dbs()
            ws = self.drivers_wb.active
            if ws.max_column != 17:  # Учитываем новую колонку ID
                logger.warning(
                    f"Ожидалось 17 колонок в базе водителей, найдено {ws.max_column}"
                )
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row[0] == driver_id:
                    logger.debug(f"Найден водитель с ID {driver_id} в строке {row_count + 1}")
                    return {
                        'ID': row[0],
                        'Водитель': row[1] or '',
                        'Паспорт_серия_и_номер': row[2] or '',
                        'Паспорт_место_выдачи': row[3] or '',
                        'Паспорт_дата_выдачи': row[4] or '',
                        'Паспорт_код_подразделения': row[5] or '',
                        'ВУ_серия_и_номер': row[6] or '',
                        'В/У_дата_срок': row[7] or '',
                        'Телефон': row[8] or '',
                        'Марка_машины_номер': row[9] or '',
                        'Прицеп_номер': row[10] or '',
                        'Перевозчик': row[11] or '',
                        'Дата_рождения': row[12] or '',
                        'Адрес_регистрации': row[13] or '',
                        'Место_рождения': row[14] or '',
                        'Место_жительства': row[15] or '',
                        'Гражданство': row[16] or ''
                    }
            logger.debug(
                f"Водитель с ID {driver_id} не найден после просмотра {row_count} строк"
            )
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске водителя с ID {driver_id}: {str(e)}")
            return None

    async def lookup_client(self, client_name):
        """
        Ищет фирму в базе данных по имени.

        Args:
            client_name (str): Название фирмы для поиска.

        Returns:
            dict: Данные фирмы, если найдена, иначе None.

        Raises:
            ValueError: Если client_name не является строкой.
            Exception: Если произошла ошибка при поиске.
        """
        if not isinstance(client_name, str):
            logger.error("client_name должен быть строкой")
            raise ValueError("client_name должен быть строкой")
        try:
            await self.load_dbs()
            ws = self.firms_wb.active
            if ws.max_column != 4:  # Учитываем новую колонку ID
                logger.warning(
                    f"Ожидалось 4 колонки в базе фирм, найдено {ws.max_column}"
                )
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row[1] and client_name.lower() in row[1].lower():
                    logger.debug(f"Найдена фирма {client_name} в строке {row_count + 1}")
                    return {
                        'ID': row[0],
                        'Краткое название': row[1] or '',
                        'Полное название': row[2] or '',
                        'ИНН': row[3] or ''
                    }
            logger.debug(
                f"Фирма {client_name} не найдена после просмотра {row_count} строк"
            )
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске фирмы {client_name}: {str(e)}")
            return None

    async def lookup_carrier(self, carrier_name):
        """
        Ищет перевозчика в базе данных по имени.

        Args:
            carrier_name (str): Название перевозчика для поиска.

        Returns:
            dict: Данные перевозчика, если найден, иначе None.

        Raises:
            ValueError: Если carrier_name не является строкой.
            Exception: Если произошла ошибка при поиске.
        """
        if not isinstance(carrier_name, str):
            logger.error("carrier_name должен быть строкой")
            raise ValueError("carrier_name должен быть строкой")
        try:
            await self.load_dbs()
            ws = self.carriers_wb.active
            logger.debug(f"Поиск перевозчика: '{carrier_name}' (длина: {len(carrier_name)})")
            logger.debug(f"Количество столбцов в базе: {ws.max_column}, ожидается 5")
            if ws.max_column != 5:  # Учитываем новую колонку ID
                logger.warning(
                    f"Ожидалось 5 колонок в базе перевозчиков, найдено {ws.max_column}"
                )
            # Логируем содержимое базы для отладки
            logger.debug("Содержимое базы перевозчиков:")
            for row in ws.iter_rows(min_row=2, values_only=True):
                logger.debug(f"Строка: {row}")
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row[1]:
                    name_in_db = row[1].strip().lower()
                    search_name = carrier_name.strip().lower()
                    logger.debug(f"Сравнение: имя в базе='{name_in_db}', искомое имя='{search_name}'")
                    if search_name in name_in_db:
                        logger.debug(f"Найден перевозчик '{carrier_name}' в строке {row_count + 1}")
                        return {
                            'ID': row[0],
                            'Краткое название': row[1] or '',
                            'Полное название': row[2] or '',
                            'ИНН': row[3] or '',
                            'Контакт': row[4] or ''
                        }
            logger.debug(
                f"Перевозчик '{carrier_name}' не найден после просмотра {row_count} строк"
            )
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске перевозчика '{carrier_name}': {str(e)}")
            return None

    async def add_driver(self, data):
        """
        Добавляет нового водителя в базу данных.

        Args:
            data (dict): Данные водителя.

        Returns:
            int: ID новой записи, если водитель добавлен, None, если уже существует.

        Raises:
            ValueError: Если data не является словарем или отсутствует ключ 'Водитель'.
            Exception: Если произошла ошибка при добавлении.
        """
        if not isinstance(data, dict):
            logger.error("data должен быть словарем")
            raise ValueError("data должен быть словарем")
        if 'Водитель' not in data:
            logger.error("Ключ 'Водитель' отсутствует в данных")
            raise ValueError("Ключ 'Водитель' отсутствует в данных")
        try:
            await self.load_dbs()
            ws = self.drivers_wb.active
            driver_exists = False
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and data['Водитель'].lower() in row[1].lower():
                    driver_exists = True
                    break
            if not driver_exists:
                new_id = ws.max_row  # ID = номер строки
                ws.append([
                    new_id,
                    data.get('Водитель', ''),
                    data.get('Паспорт_серия_и_номер', ''),
                    data.get('Паспорт_место_выдачи', ''),
                    data.get('Паспорт_дата_выдачи', ''),
                    data.get('Паспорт_код_подразделения', ''),
                    data.get('ВУ_серия_и_номер', ''),
                    data.get('В/У_дата_срок', ''),
                    data.get('Телефон', ''),
                    data.get('Марка_машины_номер', ''),
                    data.get('Прицеп_номер', ''),
                    data.get('Перевозчик', ''),
                    data.get('Дата_рождения', ''),
                    data.get('Адрес_регистрации', ''),
                    data.get('Место_рождения', ''),
                    data.get('Место_жительства', ''),
                    data.get('Гражданство', '')
                ])
                self.drivers_modified = True  # Устанавливаем флаг
                await self.save_dbs()
                logger.info(f"Добавлен новый водитель в базу: {data['Водитель']} с ID {new_id}")
                return new_id
            logger.info(f"Водитель уже существует в базе: {data['Водитель']}")
            return None
        except Exception as e:
            logger.error(f"Ошибка при добавлении водителя в базу: {str(e)}")
            return None

    async def update_driver(self, data):
        """
        Обновляет данные существующего водителя в базе.

        Args:
            data (dict): Обновлённые данные водителя, включая 'ID'.

        Returns:
            bool: True, если данные обновлены, False, если водитель не найден.

        Raises:
            ValueError: Если data не является словарем или отсутствует ключ 'ID'.
            Exception: Если произошла ошибка при обновлении.
        """
        if not isinstance(data, dict):
            logger.error("data должен быть словарем")
            raise ValueError("data должен быть словарем")
        if 'ID' not in data:
            logger.error("Ключ 'ID' отсутствует в данных")
            raise ValueError("Ключ 'ID' отсутствует в данных")
        try:
            await self.load_dbs()
            ws = self.drivers_wb.active
            for row in ws.iter_rows(min_row=2):
                if row[0].value == data['ID']:
                    row[1].value = data.get('Водитель', '')
                    row[2].value = data.get('Паспорт_серия_и_номер', '')
                    row[3].value = data.get('Паспорт_место_выдачи', '')
                    row[4].value = data.get('Паспорт_дата_выдачи', '')
                    row[5].value = data.get('Паспорт_код_подразделения', '')
                    row[6].value = data.get('ВУ_серия_и_номер', '')
                    row[7].value = data.get('В/У_дата_срок', '')
                    row[8].value = data.get('Телефон', '')
                    row[9].value = data.get('Марка_машины_номер', '')
                    row[10].value = data.get('Прицеп_номер', '')
                    row[11].value = data.get('Перевозчик', '')
                    row[12].value = data.get('Дата_рождения', '')
                    row[13].value = data.get('Адрес_регистрации', '')
                    row[14].value = data.get('Место_рождения', '')
                    row[15].value = data.get('Место_жительства', '')
                    row[16].value = data.get('Гражданство', '')
                    self.drivers_modified = True  # Устанавливаем флаг
                    await self.save_dbs()
                    logger.info(f"Данные водителя с ID {data['ID']} обновлены в базе")
                    return True
            logger.warning(f"Водитель с ID {data['ID']} не найден для обновления")
            return False
        except Exception as e:
            logger.error(f"Ошибка при обновлении данных водителя с ID {data['ID']}: {str(e)}")
            return False

    async def add_client(self, data):
        """
        Добавляет новую фирму в базу данных.

        Args:
            data (dict): Данные фирмы.

        Returns:
            int: ID новой записи, если фирма добавлена, None, если уже существует.

        Raises:
            ValueError: Если data не является словарем или отсутствует ключ 'Краткое название'.
            Exception: Если произошла ошибка при добавлении.
        """
        if not isinstance(data, dict):
            logger.error("data должен быть словарем")
            raise ValueError("data должен быть словарем")
        if 'Краткое название' not in data:
            logger.error("Ключ 'Краткое название' отсутствует в данных")
            raise ValueError("Ключ 'Краткое название' отсутствует в данных")
        try:
            await self.load_dbs()
            ws = self.firms_wb.active
            client_exists = False
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and data['Краткое название'].lower() in row[1].lower():
                    client_exists = True
                    break
            if not client_exists:
                new_id = ws.max_row  # ID = номер строки
                ws.append([
                    new_id,
                    data.get('Краткое название', ''),
                    data.get('Полное название', ''),
                    data.get('ИНН', '')
                ])
                self.firms_modified = True  # Устанавливаем флаг
                await self.save_dbs()
                logger.info(f"Добавлена новая фирма в базу: {data['Краткое название']} с ID {new_id}")
                return new_id
            logger.info(f"Фирма уже существует в базе: {data['Краткое название']}")
            return None
        except Exception as e:
            logger.error(f"Ошибка при добавлении фирмы в базу: {str(e)}")
            return None

    async def add_carrier(self, data):
        """
        Добавляет нового перевозчика в базу данных.

        Args:
            data (dict): Данные перевозчика.

        Returns:
            int: ID новой записи, если перевозчик добавлен, None, если уже существует.

        Raises:
            ValueError: Если data не является словарем или отсутствует ключ 'Краткое название'.
            Exception: Если произошла ошибка при добавлении.
        """
        if not isinstance(data, dict):
            logger.error("data должен быть словарем")
            raise ValueError("data должен быть словарем")
        if 'Краткое название' not in data:
            logger.error("Ключ 'Краткое название' отсутствует в данных")
            raise ValueError("Ключ 'Краткое название' отсутствует в данных")
        try:
            await self.load_dbs()
            ws = self.carriers_wb.active
            carrier_exists = False
            normalized_name = data['Краткое название'].strip().lower()
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and normalized_name in row[1].strip().lower():
                    carrier_exists = True
                    break
            if not carrier_exists:
                new_id = ws.max_row  # ID = номер строки
                ws.append([
                    new_id,
                    data.get('Краткое название', '').strip(),
                    data.get('Полное название', '').strip(),
                    data.get('ИНН', '').strip(),
                    data.get('Контакт', '').strip()
                ])
                self.carriers_modified = True  # Устанавливаем флаг изменения
                await self.save_dbs()
                logger.info(f"Добавлен новый перевозчик в базу: {data['Краткое название']} с ID {new_id}")
                return new_id
            logger.info(f"Перевозчик уже существует в базе: {data['Краткое название']}")
            return None
        except Exception as e:
            logger.error(f"Ошибка при добавлении перевозчика в базу: {str(e)}")
            return None

    async def update_carrier(self, data):
        """
        Обновляет данные существующего перевозчика в базе.

        Args:
            data (dict): Обновлённые данные перевозчика, включая 'ID'.

        Returns:
            bool: True, если данные обновлены, False, если перевозчик не найден.

        Raises:
            ValueError: Если data не является словарем или отсутствует ключ 'ID'.
            Exception: Если произошла ошибка при обновлении.
        """
        if not isinstance(data, dict):
            logger.error("data должен быть словарем")
            raise ValueError("data должен быть словарем")
        if 'ID' not in data:
            logger.error("Ключ 'ID' отсутствует в данных")
            raise ValueError("Ключ 'ID' отсутствует в данных")
        try:
            await self.load_dbs()
            ws = self.carriers_wb.active
            for row in ws.iter_rows(min_row=2):
                if row[0].value == data['ID']:
                    row[1].value = data.get('Краткое название', '').strip()
                    row[2].value = data.get('Полное название', '').strip()
                    row[3].value = data.get('ИНН', '').strip()
                    row[4].value = data.get('Контакт', '').strip()
                    self.carriers_modified = True  # Устанавливаем флаг изменения
                    await self.save_dbs()
                    logger.info(f"Данные перевозчика с ID {data['ID']} обновлены в базе")
                    return True
            logger.warning(f"Перевозчик с ID {data['ID']} не найден для обновления")
            return False
        except Exception as e:
            logger.error(f"Ошибка при обновлении данных перевозчика с ID {data['ID']}: {str(e)}")
            return False

class YandexDiskUploader:
    """Класс для загрузки файлов на Яндекс.Диск."""

    def __init__(self, y_disk):
        """
        Инициализирует объект YandexDiskUploader.

        Args:
            y_disk (yadisk.AsyncYaDisk): Клиент Яндекс.Диска.

        Raises:
            ValueError: Если y_disk не является объектом yadisk.AsyncYaDisk.
        """
        if not isinstance(y_disk, yadisk.AsyncYaDisk):
            logger.error("y_disk должен быть объектом yadisk.AsyncYaDisk")
            raise ValueError("y_disk должен быть объектом yadisk.AsyncYaDisk")
        self.y_disk = y_disk
        logger.info("YandexDiskUploader инициализирован")

    async def ensure_path_exists(self, path):
        """
        Проверяет и создаёт путь на Яндекс.Диске, если он не существует.

        Args:
            path (str): Путь на Яндекс.Диске.

        Raises:
            ValueError: Если path не является строкой.
            Exception: Если не удалось создать директорию.
        """
        if not isinstance(path, str):
            logger.error("path должен быть строкой")
            raise ValueError("path должен быть строкой")
        parts = path.strip('/').split('/')
        current_path = ''
        for part in parts:
            if part:
                current_path += f'/{part}'
                exists = await self.y_disk.exists(current_path)
                if not exists:
                    try:
                        await self.y_disk.mkdir(current_path)
                        logger.info(f"Создана директория: {current_path}")
                    except Exception as e:
                        logger.error(f"Ошибка создания директории {current_path}: {str(e)}")
                        raise

    async def upload_files(self, files, message, retries=10, delay=15):
        """
        Загружает файлы на Яндекс.Диск с несколькими попытками.

        Args:
            files (list): Список путей к локальным файлам.
            message: Сообщение Telegram для отправки уведомлений.
            retries (int): Количество попыток загрузки.
            delay (int): Задержка между попытками в секундах.

        Returns:
            bool: True, если загрузка успешна, False в случае ошибки.
        """
        current_month = datetime.now().strftime("%Y_%m")
        remote_path = f"/TransportData/{current_month}"
        try:
            await self.ensure_path_exists(remote_path)
        except Exception as e:
            logger.error(f"Ошибка структуры: {str(e)}")
            await bot.send_message(
                chat_id=message.chat.id,
                text="⚠️ Ошибка создания структуры на Яндекс.Диск."
            )
            return False

        for file_path in files:
            for attempt in range(retries):
                try:
                    if not os.path.exists(file_path):
                        logger.error(f"Файл не найден: {file_path}")
                        continue
                    filename = os.path.basename(file_path)
                    remote_file = f"{remote_path}/{filename}"
                    logger.debug(f"Попытка загрузки файла {filename} на Яндекс.Диск")
                    await self.y_disk.upload(file_path, remote_file, overwrite=True)
                    logger.info(f"✅ {filename} загружен")
                    await bot.send_message(
                        chat_id=message.chat.id,
                        text=f"📤 Файл {filename} успешно загружен на Яндекс.Диск!"
                    )
                    break
                except Exception as e:
                    logger.error(
                        f"Ошибка загрузки {filename} "
                        f"(попытка {attempt + 1}/{retries}): {str(e)}"
                    )
                    if attempt < retries - 1:
                        await asyncio.sleep(delay)
                    else:
                        logger.error(f"Не удалось загрузить {filename} после всех попыток")
                        await bot.send_message(
                            chat_id=message.chat.id,
                            text=(
                                f"⚠️ Ошибка загрузки файла {filename} на Яндекс.Диск. "
                                f"Данные сохранены локально. Попробуйте загрузить вручную "
                                f"из {file_path} позже."
                            )
                        )
                        return False
        return True

class ExcelManager:
    """Класс для работы с Excel-файлами (ежедневник, фирмы, перевозчики)."""

    def __init__(self, base_dir):
        """
        Инициализирует объект ExcelManager.

        Args:
            base_dir (str): Базовая директория для хранения файлов.

        Raises:
            ValueError: Если base_dir не является строкой.
        """
        if not isinstance(base_dir, str):
            logger.error("base_dir должен быть строкой")
            raise ValueError("base_dir должен быть строкой")
        if not os.path.exists(base_dir):
            os.makedirs(base_dir, exist_ok=True)
            logger.info(f"Создана базовая директория: {base_dir}")
        self.base_dir = base_dir
        self.current_month = datetime.now().strftime("%Y_%m")
        self.column_widths = {
            'daily': [15, 25, 20, 50, 25, 15, 15, 15, 20],
            'firms': [15, 25, 20, 50, 15, 20],
            'carriers': [15, 25, 25, 20, 50, 15, 20]
        }
        self.colors = {
            'header': 'FF4F81BD',
            'even_row': 'FFFFFFFF',
            'odd_row': 'FFF2F2F2'
        }
        logger.info("ExcelManager инициализирован")

    def get_file_paths(self):
        """
        Возвращает пути к файлам Excel для ежедневника, фирм и перевозчиков.

        Returns:
            dict: Словарь с путями к файлам.
        """
        return {
            'daily': os.path.join(self.base_dir, f'Ежедневник_{self.current_month}.xlsx'),
            'firms': os.path.join(
                self.base_dir,
                f'Фирмы-заказчики_{self.current_month}.xlsx'
            ),
            'carriers': os.path.join(
                self.base_dir,
                f'Перевозчики_{self.current_month}.xlsx'
            )
        }

    def _set_column_widths(self, sheet, widths):
        """
        Устанавливает ширину столбцов в таблице Excel.

        Args:
            sheet: Лист Excel.
            widths (list): Список ширин столбцов.
        """
        for col_idx, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width + 2

    def _apply_header_style(self, cell):
        """
        Применяет стиль к заголовкам в таблице Excel.

        Args:
            cell: Ячейка Excel.
        """
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(
            start_color=self.colors['header'],
            end_color=self.colors['header'],
            fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='medium', color='FF666666'),
            right=Side(style='medium', color='FF666666'),
            top=Side(style='medium', color='FF666666'),
            bottom=Side(style='medium', color='FF666666')
        )

    def _apply_cell_style(self, cell, row_num, align='left', number_format=None):
        """
        Применяет стиль к ячейкам данных в таблице Excel.

        Args:
            cell: Ячейка Excel.
            row_num (int): Номер строки (для чередования цветов).
            align (str): Выравнивание текста (по умолчанию 'left').
            number_format (str): Формат числа (если требуется).
        """
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        color = self.colors['even_row'] if row_num % 2 == 0 else self.colors['odd_row']
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.border = Border(
            left=Side(style='thin', color='FF666666'),
            right=Side(style='thin', color='FF666666'),
            top=Side(style='thin', color='FF666666'),
            bottom=Side(style='thin', color='FF666666')
        )
        if number_format:
            cell.number_format = number_format

    def ensure_files_exist(self):
        """
        Проверяет существование Excel-файлов и создаёт их при необходимости.
        """
        files = self.get_file_paths()
        headers = {
            'daily': [
                'Дата', 'Фирма', 'Направление', 'Водитель', 'Перевозчик',
                'Цена', 'Оплата', 'Разница', 'Пометка'
            ],
            'firms': ['Дата', 'Фирма', 'Направление', 'Водитель', 'Цена', 'Пометка'],
            'carriers': [
                'Дата', 'Перевозчик', 'Фирма', 'Направление',
                'Водитель', 'Оплата', 'Пометка'
            ]
        }
        for file_type, path in files.items():
            if not os.path.exists(path):
                logger.info(f"Создание нового файла: {path}")
                wb = Workbook()
                ws = wb.active
                ws.title = self.current_month
                self._set_column_widths(ws, self.column_widths[file_type])
                for col_idx, header in enumerate(headers[file_type], 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    self._apply_header_style(cell)
                wb.save(path)
                logger.info(f"Создан новый файл: {path}")

    def add_record(self, file_path, data, file_type):
        """
        Добавляет запись в Excel-файл.

        Args:
            file_path (str): Путь к файлу Excel.
            data (dict): Данные для записи.
            file_type (str): Тип файла ('daily', 'firms', 'carriers').

        Returns:
            bool: True, если запись добавлена, False в случае ошибки.
        """
        try:
            logger.debug(f"Добавление записи в файл: {file_path}, тип: {file_type}")
            wb = load_workbook(file_path)
            ws = wb.active
            row_num = ws.max_row + 1

            if file_type == 'daily':
                difference = (data.get('Цена', 0) or 0) - (data.get('Оплата', 0) or 0)
                row_data = [
                    data.get('Дата_перевозки', datetime.now().strftime('%d.%m.%Y')),
                    data.get('Фирма', 'Не указана'),
                    data.get('Направление', 'Не указано'),
                    data.get('Водитель', 'Не указан'),
                    data.get('Перевозчик', 'Не указан'),
                    data.get('Цена', 0),
                    data.get('Оплата', 0),
                    difference,
                    data.get('Пометка', 'Нет')
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    align = 'center' if col_idx in [1, 6, 7, 8] else 'left'
                    number_format = '#,##0' if col_idx in [6, 7, 8] else None
                    self._apply_cell_style(
                        cell,
                        row_num,
                        align=align,
                        number_format=number_format
                    )

            elif file_type == 'firms':
                row_data = [
                    data.get('Дата_перевозки', datetime.now().strftime('%d.%m.%Y')),
                    data.get('Фирма', 'Не указана'),
                    data.get('Направление', 'Не указано'),
                    data.get('Водитель', 'Не указан'),
                    data.get('Цена', 0),
                    data.get('Пометка', 'Нет')
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    align = 'center' if col_idx in [1, 5] else 'left'
                    number_format = '#,##0' if col_idx == 5 else None
                    self._apply_cell_style(
                        cell,
                        row_num,
                        align=align,
                        number_format=number_format
                    )

            elif file_type == 'carriers':
                row_data = [
                    data.get('Дата_перевозки', datetime.now().strftime('%d.%m.%Y')),
                    data.get('Перевозчик', 'Не указан'),
                    data.get('Фирма', 'Не указана'),
                    data.get('Направление', 'Не указано'),
                    data.get('Водитель', 'Не указан'),
                    data.get('Оплата', 0),
                    data.get('Пометка', 'Нет')
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    align = 'center' if col_idx in [1, 6] else 'left'
                    number_format = '#,##0' if col_idx == 6 else None
                    self._apply_cell_style(
                        cell,
                        row_num,
                        align=align,
                        number_format=number_format
                    )

            wb.save(file_path)
            logger.info(f"Добавлена запись в {file_path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при добавлении записи в {file_path}: {str(e)}")
            return False

# Часть 4 завершена




# Часть 5: Инициализация объектов и клавиатуры

from aiogram.fsm.state import State, StatesGroup

# Определение состояний
class Form(StatesGroup):
    add_driver = State()
    add_client = State()
    add_carrier = State()
    confirm_driver = State()
    confirm_client = State()
    confirm_carrier = State()
    add_transportation = State()  # Добавлено состояние для добавления перевозки
    confirm_transportation = State()  # Добавлено состояние для подтверждения перевозки

class DriverStates(StatesGroup):
    waiting_for_update_confirmation = State()
    waiting_for_add_confirmation = State()
    waiting_for_carrier_update_confirmation = State()

logger.info("Состояния Form и DriverStates инициализированы")

# Инициализация объектов
db_manager = DatabaseManager(y_disk=y_disk, local_dir=LOCAL_TEMP_DIR)
uploader = YandexDiskUploader(y_disk=y_disk)
excel_manager = ExcelManager(base_dir=BASE_DIR)
logger.info("Объекты db_manager, uploader, excel_manager инициализированы")

# Клавиатуры
def create_main_menu():
    """
    Создаёт главное меню с основными действиями.

    Returns:
        ReplyKeyboardMarkup: Клавиатура главного меню.
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Запись")],
            [KeyboardButton(text="➕ Добавить")],
            [KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    logger.debug("Создано главное меню")
    return keyboard

def create_record_submenu():
    """
    Создаёт подменю для ввода данных с кнопкой 'Назад'.

    Returns:
        ReplyKeyboardMarkup: Клавиатура подменю.
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    logger.debug("Создано подменю для ввода данных")
    return keyboard

def create_add_submenu():
    """
    Создаёт подменю для добавления водителя, фирмы или перевозчика.

    Returns:
        ReplyKeyboardMarkup: Клавиатура подменю.
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👤 Добавить водителя"),
             KeyboardButton(text="🏢 Добавить фирму")],
            [KeyboardButton(text="🚚 Добавить перевозчика"),
             KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    logger.debug("Создано подменю для добавления")
    return keyboard

def create_clients_submenu():
    """
    Создаёт подменю для управления клиентами.

    Returns:
        ReplyKeyboardMarkup: Клавиатура подменю.
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Добавить клиента"),
             KeyboardButton(text="📋 Просмотреть клиентов")],
            [KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    logger.debug("Создано подменю для клиентов")
    return keyboard

def create_carriers_submenu():
    """
    Создаёт подменю для управления перевозчиками.

    Returns:
        ReplyKeyboardMarkup: Клавиатура подменю.
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Добавить перевозчика"),
             KeyboardButton(text="📋 Просмотреть перевозчиков")],
            [KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    logger.debug("Создано подменю для перевозчиков")
    return keyboard

def get_update_keyboard():
    """
    Создаёт инлайн-клавиатуру для подтверждения или отмены действия.

    Returns:
        InlineKeyboardMarkup: Инлайн-клавиатура.
    """
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Обновить данные", callback_data="update_driver")],
        [InlineKeyboardButton(text="❌ Пропустить", callback_data="skip_update")]
    ])
    logger.debug("Создана инлайн-клавиатура для обновления данных")
    return keyboard

# Часть 5 завершена



# Часть 6: Хендлеры (до process_add_carrier включительно)

@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    """
    Обрабатывает команду /start, показывая главное меню.

    Args:
        message (types.Message): Входящее сообщение.
    """
    try:
        await message.answer(
            "🚛 Добро пожаловать в бот для учета перевозок!\n\n"
            "Я помогу вам вести учет водителей, фирм и перевозок. Выберите действие:",
            reply_markup=create_main_menu()
        )
        logger.info(f"Пользователь {message.from_user.id} запустил бота")
    except Exception as e:
        logger.error(f"Ошибка при обработке команды /start: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.")

@dp.message(lambda message: message.text == "➕ Добавить")
async def manage_additions(message: types.Message):
    """
    Показывает подменю для добавления водителя, фирмы или перевозчика.

    Args:
        message (types.Message): Входящее сообщение.
    """
    try:
        await message.answer(
            "➕ Выберите, кого добавить:",
            reply_markup=create_add_submenu()
        )
        logger.info(f"Пользователь {message.from_user.id} перешёл в меню добавления")
    except Exception as e:
        logger.error(f"Ошибка при переходе в меню добавления: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(lambda message: message.text == "👤 Добавить водителя")
async def add_driver(message: types.Message, state: FSMContext):
    """
    Запрашивает данные водителя и переходит в состояние добавления водителя.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        await message.answer(
            "📝 Введите данные водителя в следующем формате:\n"
            "Водитель: [ФИО]\n"
            "Паспорт: [серия и номер] выдан [место выдачи] [дата выдачи]\n"
            "В/У: [серия и номер] выдан [дата выдачи]\n"
            "Дата рождения: [дд.мм.гггг]\n"
            "Место рождения: [место рождения]\n"
            "Гражданство: [гражданство]\n"
            "Место жительства: [адрес]\n"
            "Телефон: [номер телефона]\n"
            "Марка машины номер: [марка] [гос. номер]\n"
            "Прицеп: [марка] [гос. номер]\n"
            "Перевозчик: [название]\n\n"
            "Пример:\n"
            "Водитель: Иванов Иван Иванович\n"
            "Паспорт: 4617 635097 выдан ТП №2 ОУФМС России 16.08.2017\n"
            "В/У: 99 27 527275 выдан 16.08.2017\n"
            "Дата рождения: 15.03.1985\n"
            "Место рождения: г. Москва\n"
            "Гражданство: РФ\n"
            "Место жительства: МО, г. Коломна, д. 11, кв. 89\n"
            "Телефон: +7 (926) 704-31-36\n"
            "Марка машины номер: DAF В 394 РО 750\n"
            "Прицеп: Schmitz ЕТ 1913 50\n"
            "Перевозчик: ИП Сидоров",
            reply_markup=create_record_submenu()
        )
        await state.set_state(Form.add_driver)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление водителя")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(lambda message: message.text == "🏢 Добавить фирму")
async def add_client(message: types.Message, state: FSMContext):
    """
    Запрашивает данные клиента (фирмы) и переходит в состояние добавления клиента.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        await message.answer(
            "📝 Введите данные фирмы в следующем формате:\n\n"
            "Полное название: [полное название]\n"
            "ИНН: [ИНН]\n",
            reply_markup=create_record_submenu()
        )
        await state.set_state(Form.add_client)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление фирмы")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(lambda message: message.text == "🚚 Добавить перевозчика")
async def add_carrier(message: types.Message, state: FSMContext):
    """
    Запрашивает данные перевозчика и переходит в состояние добавления перевозчика.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        await message.answer(
            "📝 Введите данные перевозчика в следующем формате:\n\n"
            "Перевозчик: [название]\n"
            "Имя: [имя, если ИП]\n"
            "Телефон: [контакт]\n"
            "ИНН: [ИНН]\n",
            reply_markup=create_record_submenu()
        )
        await state.set_state(Form.add_carrier)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление перевозчика")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(lambda message: message.text == "📋 Просмотреть клиентов")
async def view_clients(message: types.Message):
    """
    Показывает список клиентов из базы данных.

    Args:
        message (types.Message): Входящее сообщение.
    """
    try:
        await db_manager.ensure_dbs_exist()
        ws = db_manager.firms_wb.active
        if ws.max_row <= 1:
            await message.answer(
                "📂 База клиентов пуста.",
                reply_markup=create_clients_submenu()
            )
            logger.info(f"Пользователь {message.from_user.id} запросил список клиентов: база пуста")
            return
        clients_list = "📋 Список клиентов:\n\n"
        for row in ws.iter_rows(min_row=2, values_only=True):
            clients_list += (
                f"ID: {row[0]}\n"
                f"Полное название: {row[2]}\n"
                f"ИНН: {row[3]}\n"
                f"{'-' * 20}\n"
            )
        await message.answer(clients_list, reply_markup=create_clients_submenu())
        logger.info(f"Пользователь {message.from_user.id} запросил список клиентов")
    except Exception as e:
        logger.error(f"Ошибка при просмотре клиентов: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при загрузке списка клиентов. Попробуйте снова.",
            reply_markup=create_clients_submenu()
        )

@dp.message(lambda message: message.text == "📋 Просмотреть перевозчиков")
async def view_carriers(message: types.Message):
    """
    Показывает список перевозчиков из базы данных.

    Args:
        message (types.Message): Входящее сообщение.
    """
    try:
        await db_manager.ensure_dbs_exist()
        ws = db_manager.carriers_wb.active
        if ws.max_row <= 1:
            await message.answer(
                "📂 База перевозчиков пуста.",
                reply_markup=create_carriers_submenu()
            )
            logger.info(f"Пользователь {message.from_user.id} запросил список перевозчиков: база пуста")
            return
        carriers_list = "📋 Список перевозчиков:\n\n"
        for row in ws.iter_rows(min_row=2, values_only=True):
            carriers_list += (
                f"ID: {row[0]}\n"
                f"Перевозчик: {row[1]}\n"
                f"Имя: {row.get('Имя перевозчика', 'Не указано')}\n"
                f"Телефон: {row[4]}\n"
                f"ИНН: {row[3]}\n"
                f"{'-' * 20}\n"
            )
        await message.answer(carriers_list, reply_markup=create_carriers_submenu())
        logger.info(f"Пользователь {message.from_user.id} запросил список перевозчиков")
    except Exception as e:
        logger.error(f"Ошибка при просмотре перевозчиков: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при загрузке списка перевозчиков. Попробуйте снова.",
            reply_markup=create_carriers_submenu()
        )

@dp.message(lambda message: message.text == "⬅️ Назад")
async def go_back(message: types.Message, state: FSMContext):
    """
    Возвращает пользователя в главное меню и очищает состояние.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        await state.clear()
        await message.answer("Вы вернулись в главное меню.", reply_markup=create_main_menu())
        logger.info(f"Пользователь {message.from_user.id} вернулся в главное меню")
    except Exception as e:
        logger.error(f"Ошибка при возвращении в главное меню: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(StateFilter(Form.add_driver))
async def process_add_driver(message: types.Message, state: FSMContext):
    """
    Обрабатывает данные водителя, проверяет наличие в базе и предлагает добавить/обновить.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        # Сохраняем текст данных водителя
        await state.update_data(driver_data=message.text)
        user_data = await state.get_data()
        text = user_data.get('driver_data', '')
        if not text:
            await message.answer(
                "Не удалось получить данные водителя. Пожалуйста, попробуйте снова."
            )
            logger.warning(f"Пустые данные водителя от пользователя {message.from_user.id}")
            return

        raw_data, data = parse_by_keywords(text, is_driver_data=True)
        if 'Водитель' not in data:
            logger.error("Поле 'Водитель' отсутствует в распарсенных данных")
            await message.answer(
                "Не удалось определить ФИО водителя. Пожалуйста, проверьте формат данных."
            )
            return

        await db_manager.ensure_dbs_exist()
        # Ищем водителя по имени, чтобы получить ID
        ws = db_manager.drivers_wb.active
        driver_id = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and data['Водитель'].lower() in row[1].lower():
                driver_id = row[0]
                break

        formatted_data = format_driver_data(data)
        if driver_id is not None:
            existing_driver = await db_manager.lookup_driver(driver_id)
            existing_formatted = format_driver_data(existing_driver)
            await message.answer(
                f"Водитель {data['Водитель']} уже есть в базе:\n\n"
                f"Текущие данные:\n{existing_formatted}\n\n"
                f"Новые данные:\n{formatted_data}\n\n"
                "Обновить данные? (Да/Нет)",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                    ],
                    resize_keyboard=True
                )
            )
            data['ID'] = driver_id  # Сохраняем ID для обновления
            await state.update_data(driver_data=data)
            await state.set_state(DriverStates.waiting_for_update_confirmation)
            logger.info(f"Водитель {data['Водитель']} найден с ID {driver_id}, запрошено обновление")
        else:
            driver_id = await db_manager.add_driver(data)
            if driver_id is not None:
                await message.answer(
                    f"Водитель {data['Водитель']} не найден в базе.\n\n"
                    f"Данные:\n{formatted_data}\n\n"
                    f"Добавить водителя? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[
                            [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                        ],
                        resize_keyboard=True
                    )
                )
                data['ID'] = driver_id  # Сохраняем ID для добавления
                await state.update_data(driver_data=data)
                await state.set_state(DriverStates.waiting_for_add_confirmation)
                logger.info(f"Водитель {data['Водитель']} не найден, запрошено добавление с ID {driver_id}")
            else:
                await message.answer(
                    "⚠️ Ошибка при добавлении водителя. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Ошибка при добавлении водителя: {data.get('Водитель', 'Неизвестно')}")
                await state.clear()

    except Exception as e:
        logger.error(f"Ошибка при обработке данных водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "Произошла ошибка при обработке данных. Пожалуйста, попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(Form.add_client))
async def process_add_client(message: types.Message, state: FSMContext):
    """
    Обрабатывает данные клиента (фирмы) и предлагает подтвердить добавление.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        text = message.text.strip()
        # Пробуем сначала разобрать данные в строгом формате
        full_name = None
        inn = None

        full_match = re.search(r"Полное название\s*:\s*(.+?)(?=\n|$)", text, re.IGNORECASE)
        inn_match = re.search(r"ИНН\s*:\s*(\d+)", text, re.IGNORECASE)

        if full_match:
            full_name = full_match.group(1).strip()
        if inn_match:
            inn = inn_match.group(1).strip()

        # Если строгий формат не сработал, пробуем свободную форму
        if not (full_name and inn):
            raw_data, data = parse_by_keywords(text, is_driver_data=False)
            full_name = data.get("Клиент") or data.get("Полное название")
            if not full_name:
                firm_match = re.search(
                    r"^(ООО|ИП|ОАО|ЗАО)\s+(.+?)(?=\s*ИНН|$)",
                    text.replace('\n', ' '),
                    re.IGNORECASE
                )
                if firm_match:
                    full_name = f"{firm_match.group(1)} {firm_match.group(2).strip()}"
                else:
                    await message.answer(
                        "Не удалось определить название фирмы. Пожалуйста, укажите данные в формате:\n"
                        "Полное название: [полное название]\n"
                        "ИНН: [ИНН]\n\n"
                        "Или в свободной форме, например:\n"
                        "ООО Мурманские Рыбопродукты\n"
                        "ИНН: 5190130170",
                        reply_markup=create_record_submenu()
                    )
                    return

            if not inn:
                inn_match = re.search(r"ИНН\s*(\d+)", text, re.IGNORECASE)
                if inn_match:
                    inn = inn_match.group(1).strip()
                else:
                    await message.answer(
                        "Не удалось определить ИНН. Пожалуйста, укажите данные в формате:\n"
                        "Полное название: [полное название]\n"
                        "ИНН: [ИНН]\n\n"
                        "Или в свободной форме, например:\n"
                        "ООО Мурманские Рыбопродукты\n"
                        "ИНН: 5190130170",
                        reply_markup=create_record_submenu()
                    )
                    return

        # Проверка длины ИНН
        full_name_lower = full_name.lower()
        if (full_name_lower.startswith("ооо") or full_name_lower.startswith("оао") or full_name_lower.startswith("зао")) and len(inn) != 10:
            await message.answer(
                "ИНН для ООО, ОАО или ЗАО должен содержать 10 цифр.",
                reply_markup=create_record_submenu()
            )
            return
        elif full_name_lower.startswith("ип") and len(inn) != 12:
            await message.answer(
                "ИНН для ИП должен содержать 12 цифр.",
                reply_markup=create_record_submenu()
            )
            return

        # Формируем данные фирмы
        client_data = {
            "Краткое название": full_name,
            "Полное название": full_name,
            "ИНН": inn
        }

        # Нормализация данных перед отображением
        normalized_client_data = normalize_data(client_data)

        # Форматируем данные для отображения (без Краткого названия)
        formatted_data = (
            f"Полное название: {normalized_client_data['Полное название']}\n"
            f"ИНН: {normalized_client_data['ИНН']}"
        )

        await message.answer(
            f"Вот как я разобрал данные фирмы:\n{formatted_data}\n\n"
            "Добавить фирму? (Да/Нет)",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                ],
                resize_keyboard=True
            )
        )
        # Ищем фирму, чтобы проверить, существует ли она
        ws = db_manager.firms_wb.active
        client_id = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and full_name.lower() in row[1].lower():
                client_id = row[0]
                break
        if client_id is not None:
            normalized_client_data['ID'] = client_id
        await state.update_data(client_data=normalized_client_data)
        await state.set_state(Form.confirm_client)
        logger.info(f"Пользователь {message.from_user.id} ввёл данные фирмы: {full_name}")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_add_submenu()
        )

@dp.message(StateFilter(Form.add_carrier))
async def process_add_carrier(message: types.Message, state: FSMContext):
    """
    Обрабатывает данные перевозчика и предлагает подтвердить добавление или обновление.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        text = message.text.strip().replace('\n', ' ')
        logger.debug(f"Обработанный текст для парсинга: {text}")

        carrier_name = None
        carrier_name_person = None
        carrier_contact = None
        inn = None

        raw_data, data = parse_by_keywords(text, is_driver_data=False)
        carrier_name = data.get("Перевозчик")
        carrier_contact = data.get("Контакт") or data.get("Телефон")

        inn_match = re.search(r"ИНН\s*(\d+)", text, re.IGNORECASE)
        if inn_match:
            inn = inn_match.group(1).strip()

        # Проверяем, удалось ли извлечь название перевозчика
        if not carrier_name:
            # Пробуем извлечь название перевозчика в формате "ООО/ИП Название"
            carrier_match = re.search(
                r"^(ООО|ИП|ОАО|ЗАО)\s+(.+?)(?=\s*(?:Имя|Телефон|ИНН|$))",
                text,
                re.IGNORECASE
            )
            if carrier_match:
                carrier_name = f"{carrier_match.group(1)} {carrier_match.group(2).strip()}"
            else:
                # Пробуем извлечь название до ИНН
                carrier_match = re.search(
                    r"^(ООО|ИП|ОАО|ЗАО)\s+(.+?)(?=\s*ИНН\s*\d+|)",
                    text,
                    re.IGNORECASE
                )
                if carrier_match:
                    carrier_name = f"{carrier_match.group(1)} {carrier_match.group(2).strip()}"

        # Если название всё ещё не найдено, проверяем наличие ФИО после "ИП"
        if not carrier_name:
            # Пробуем извлечь "ИП Фамилия Имя Отчество"
            ip_match = re.search(
                r"^(ИП)\s+([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)",
                text,
                re.IGNORECASE
            )
            if ip_match:
                carrier_name = f"{ip_match.group(1)} {ip_match.group(2).strip()}"
            else:
                await message.answer(
                    "Не удалось определить название перевозчика. Пожалуйста, укажите данные в формате:\n"
                    "Перевозчик: [название]\n"
                    "Имя: [имя, если ИП]\n"
                    "Телефон: [контакт]\n"
                    "ИНН: [ИНН]\n\n"
                    "Или в свободной форме, например:\n"
                    "ИП Помидоров Иван Иванович Телефон +7 (123) 456-78-90 ИНН 123456789012",
                    reply_markup=create_record_submenu()
                )
                return

        # Извлекаем ФИО, если это ИП
        if carrier_name.lower().startswith("ип"):
            name_match = re.search(
                r"ип\s+([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)",
                carrier_name,
                re.IGNORECASE
            )
            if name_match:
                carrier_name_person = name_match.group(1).strip()
            else:
                # Пробуем извлечь ФИО после ИНН или телефона
                name_after_text_match = re.search(
                    r"(?:ИНН\s*\d+\s*|\+?\d\s*\(?\d{3}\)?\s*\d{3}\-?\d{2}\-?\d{2}\s*|\d\s*\d{3}\s*\d{3}\d{2}\d{2}\s*)(\b[А-ЯЁ][а-яё]+\b\s+[А-ЯЁ][а-яё]+\b\s+[А-ЯЁ][а-яё]+\b|\b[А-ЯЁ][а-яё]+\b)",
                    text,
                    re.IGNORECASE
                )
                if name_after_text_match:
                    carrier_name_person = name_after_text_match.group(1).strip()

        # Извлекаем ИНН, если ещё не найден
        if not inn:
            inn_match = re.search(r"ИНН\s*(\d+)", text, re.IGNORECASE)
            if inn_match:
                inn = inn_match.group(1).strip()

        # Извлекаем контакт, если ещё не найден
        if not carrier_contact:
            contact_match = re.search(
                r"(?:Телефон\s*:\s*|Телефон\s+)?(\+?\d\s*\(?\d{3}\)?\s*\d{3}\-?\d{2}\-?\d{2}|\d\s*\d{3}\s*\d{3}\d{2}\d{2})",
                text,
                re.IGNORECASE
            )
            if contact_match:
                contact_text = contact_match.group(1).strip()
                carrier_contact = parse_phone_numbers(contact_text)

        # Проверяем, все ли обязательные поля заполнены
        if not carrier_name:
            await message.answer(
                "Не удалось определить название перевозчика. Пожалуйста, укажите данные в формате:\n"
                "Перевозчик: [название]\n"
                "Имя: [имя, если ИП]\n"
                "Телефон: [контакт]\n"
                "ИНН: [ИНН]\n\n"
                "Или в свободной форме, например:\n"
                "ИП Помидоров Иван Иванович Телефон +7 (123) 456-78-90 ИНН 123456789012",
                reply_markup=create_record_submenu()
            )
            return

        if not inn:
            await message.answer(
                "Не удалось определить ИНН. Пожалуйста, укажите данные в формате:\n"
                "Перевозчик: [название]\n"
                "Имя: [имя, если ИП]\n"
                "Телефон: [контакт]\n"
                "ИНН: [ИНН]\n\n"
                "Или в свободной форме, например:\n"
                "ИП Помидоров Иван Иванович Телефон +7 (123) 456-78-90 ИНН 123456789012",
                reply_markup=create_record_submenu()
            )
            return

        if not carrier_contact:
            await message.answer(
                "Некорректный формат телефона или телефон не указан. Укажите номер телефона, например: +7 (123) 456-78-90",
                reply_markup=create_record_submenu()
            )
            return

        # Проверяем длину ИНН
        carrier_name_lower = carrier_name.lower()
        if carrier_name_lower.startswith("ип") and len(inn) != 12:
            await message.answer(
                "ИНН для ИП должен содержать 12 цифр.",
                reply_markup=create_record_submenu()
            )
            return
        elif (carrier_name_lower.startswith("ооо") or carrier_name_lower.startswith("оао") or carrier_name_lower.startswith("зао")) and len(inn) != 10:
            await message.answer(
                "ИНН для ООО, ОАО или ЗАО должен содержать 10 цифр.",
                reply_markup=create_record_submenu()
            )
            return

        # Формируем данные перевозчика
        carrier_data = {
            "Краткое название": carrier_name,
            "Полное название": carrier_name,
            "ИНН": inn,
            "Контакт": carrier_contact
        }
        if carrier_name_person:
            carrier_data["Имя перевозчика"] = carrier_name_person

        normalized_carrier_data = normalize_data(carrier_data)
        formatted_data = (
            f"Перевозчик: {normalized_carrier_data['Краткое название']}\n"
            f"Имя: {normalized_carrier_data.get('Имя перевозчика', 'Не указано')}\n"
            f"Телефон: {normalized_carrier_data['Контакт']}\n"
            f"ИНН: {normalized_carrier_data['ИНН']}"
        )

        # Проверяем, существует ли перевозчик в базе
        await db_manager.ensure_dbs_exist()
        ws = db_manager.carriers_wb.active
        # Логируем содержимое базы для отладки
        logger.debug(f"Содержимое базы перевозчиков (carriers_db.xlsx):")
        for row in ws.iter_rows(min_row=2, values_only=True):
            logger.debug(f"Строка: {row}")
        carrier_id = None
        existing_carrier = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and carrier_name.lower() in row[1].lower():
                carrier_id = row[0]
                break

        if carrier_id is not None:
            existing_carrier = await db_manager.lookup_carrier(carrier_name)
            if existing_carrier:
                existing_formatted = (
                    f"Перевозчик: {existing_carrier['Краткое название']}\n"
                    f"Имя: {existing_carrier.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {existing_carrier['Контакт']}\n"
                    f"ИНН: {existing_carrier['ИНН']}"
                )
                await message.answer(
                    f"Перевозчик {carrier_name} уже есть в базе:\n\n"
                    f"Текущие данные:\n{existing_formatted}\n\n"
                    f"Новые данные:\n{formatted_data}\n\n"
                    "Обновить данные? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[
                            [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                        ],
                        resize_keyboard=True
                    )
                )
                normalized_carrier_data['ID'] = carrier_id
                await state.update_data(carrier_data=normalized_carrier_data)
                await state.set_state(DriverStates.waiting_for_carrier_update_confirmation)
                logger.info(f"Перевозчик {carrier_name} найден с ID {carrier_id}, запрошено обновление")
            else:
                # Добавляем логирование для отладки
                logger.error(f"Не удалось получить данные перевозчика с ID {carrier_id}")
                # Выводим сообщение о том, что перевозчик уже есть, даже если данные не получены
                await message.answer(
                    f"⚠️ Перевозчик {carrier_name} уже существует в базе, но не удалось получить его данные.\n"
                    f"Новые данные:\n{formatted_data}\n\n"
                    "Обновить данные? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[
                            [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                        ],
                        resize_keyboard=True
                    )
                )
                normalized_carrier_data['ID'] = carrier_id
                await state.update_data(carrier_data=normalized_carrier_data)
                await state.set_state(DriverStates.waiting_for_carrier_update_confirmation)
        else:
            await message.answer(
                f"Вот как я разобрал данные перевозчика:\n{formatted_data}\n\n"
                "Добавить перевозчика? (Да/Нет)",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                    ],
                    resize_keyboard=True
                )
            )
            await state.update_data(carrier_data=normalized_carrier_data)
            await state.set_state(Form.confirm_carrier)
            logger.info(f"Пользователь {message.from_user.id} ввёл данные перевозчика: {carrier_name}")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_add_submenu()
        )
        await state.clear()

# Часть 6 завершена



# Часть 7: Оставшиеся хендлеры и запуск бота

@dp.message(StateFilter(Form.confirm_client))
async def confirm_add_client(message: types.Message, state: FSMContext):
    """
    Подтверждает добавление клиента (фирмы) в базу данных.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_client, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        client_data = user_data.get("client_data", {})
        if not client_data:
            logger.error("client_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные фирмы не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            if 'ID' in client_data:
                # Фирма уже существует, обновление не предусмотрено, сообщаем об этом
                await message.answer(
                    "⚠️ Фирма уже существует в базе. Обновление данных фирмы не предусмотрено.",
                    reply_markup=create_main_menu()
                )
                logger.info(f"Пользователь {message.from_user.id} попытался добавить существующую фирму: {client_data['Полное название']}")
            else:
                client_id = await db_manager.add_client(client_data)
                if client_id is not None:
                    formatted_data = (
                        f"Полное название: {client_data['Полное название']}\n"
                        f"ИНН: {client_data['ИНН']}"
                    )
                    await message.answer(
                        f"✅ Фирма успешно добавлена с ID {client_id}:\n\n{formatted_data}",
                        reply_markup=create_main_menu()
                    )
                    logger.info(f"Фирма {client_data['Полное название']} успешно добавлена с ID {client_id}")
                else:
                    await message.answer(
                        "⚠️ Ошибка при добавлении фирмы. Возможно, фирма с таким названием уже существует.",
                        reply_markup=create_main_menu()
                    )
                    logger.error(f"Ошибка при добавлении фирмы {client_data.get('Полное название', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Добавление фирмы отменено.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил добавление фирмы")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении фирмы. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(Form.confirm_carrier))
async def confirm_add_carrier(message: types.Message, state: FSMContext):
    """
    Подтверждает добавление перевозчика в базу данных.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_carrier, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        carrier_data = user_data.get("carrier_data", {})
        carrier_name = carrier_data.get("Краткое название", "Неизвестно")
        if not carrier_data:
            logger.error("carrier_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозчика не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            # Проверяем, существует ли перевозчик в базе
            existing_carrier = await db_manager.lookup_carrier(carrier_name)
            if existing_carrier:
                await message.answer(
                    f"⚠️ Перевозчик {carrier_name} уже существует в базе:\n"
                    f"Перевозчик: {existing_carrier['Краткое название']}\n"
                    f"Имя: {existing_carrier.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {existing_carrier['Контакт']}\n"
                    f"ИНН: {existing_carrier['ИНН']}\n\n"
                    "Чтобы обновить данные, повторите ввод с изменёнными данными.",
                    reply_markup=create_main_menu()
                )
                logger.info(f"Попытка добавить уже существующего перевозчика: {carrier_name}")
            else:
                carrier_id = await db_manager.add_carrier(carrier_data)
                if carrier_id is not None:
                    formatted_data = (
                        f"Перевозчик: {carrier_data['Краткое название']}\n"
                        f"Имя: {carrier_data.get('Имя перевозчика', 'Не указано')}\n"
                        f"Телефон: {carrier_data['Контакт']}\n"
                        f"ИНН: {carrier_data['ИНН']}"
                    )
                    await message.answer(
                        f"✅ Перевозчик успешно добавлен с ID {carrier_id}:\n\n{formatted_data}",
                        reply_markup=create_main_menu()
                    )
                    logger.info(f"Перевозчик {carrier_data['Краткое название']} успешно добавлен с ID {carrier_id}")
                else:
                    await message.answer(
                        "⚠️ Ошибка при добавлении перевозчика. Возможно, перевозчик с таким названием уже существует.",
                        reply_markup=create_main_menu()
                    )
                    logger.error(f"Ошибка при добавлении перевозчика {carrier_data.get('Краткое название', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Добавление перевозчика отменено.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил добавление перевозчика")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении перевозчика. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(DriverStates.waiting_for_carrier_update_confirmation))
async def confirm_update_carrier(message: types.Message, state: FSMContext):
    """
    Подтверждает обновление данных перевозчика в базе.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_update_carrier, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        carrier_data = user_data.get("carrier_data", {})
        carrier_name = carrier_data.get("Краткое название", "Неизвестно")
        if not carrier_data:
            logger.error("carrier_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозчика не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if 'ID' not in carrier_data:
            logger.error("ID перевозчика отсутствует в данных")
            await message.answer(
                "⚠️ ID перевозчика не найден. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            success = await db_manager.update_carrier(carrier_data)
            if success:
                formatted_data = (
                    f"Перевозчик: {carrier_data['Краткое название']}\n"
                    f"Имя: {carrier_data.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {carrier_data['Контакт']}\n"
                    f"ИНН: {carrier_data['ИНН']}"
                )
                await message.answer(
                    f"✅ Данные перевозчика с ID {carrier_data['ID']} обновлены:\n\n{formatted_data}",
                    reply_markup=create_main_menu()
                )
                logger.info(f"Данные перевозчика {carrier_data['Краткое название']} с ID {carrier_data['ID']} обновлены")
            else:
                await message.answer(
                    f"⚠️ Не удалось обновить данные перевозчика {carrier_name}. Возможно, перевозчик с ID {carrier_data['ID']} не найден.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Ошибка при обновлении данных перевозчика с ID {carrier_data.get('ID', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Обновление данных перевозчика отменено.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил обновление данных перевозчика")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении обновления перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обновлении данных перевозчика. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(DriverStates.waiting_for_add_confirmation))
async def confirm_add_driver(message: types.Message, state: FSMContext):
    """
    Обрабатывает подтверждение добавления водителя.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_driver, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("driver_data", {})
        if not data:
            logger.error("driver_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные водителя не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            # Водитель уже добавлен в базе на этапе process_add_driver, просто подтверждаем
            await message.answer(
                f"✅ Водитель успешно добавлен с ID {data['ID']}:\n\n{format_driver_data(data)}",
                reply_markup=create_main_menu()
            )
            logger.info(f"Водитель {data.get('Водитель', 'Неизвестно')} успешно подтверждён с ID {data['ID']}")
        else:
            # Здесь можно добавить логику удаления, если нужно
            await message.answer(
                "⏭ Добавление водителя отменено.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил добавление водителя")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении водителя. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(DriverStates.waiting_for_update_confirmation))
async def confirm_update_driver(message: types.Message, state: FSMContext):
    """
    Обрабатывает подтверждение обновления данных водителя.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_update_driver, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("driver_data", {})
        if not data:
            logger.error("driver_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные водителя не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            success = await db_manager.update_driver(data)
            if success:
                await message.answer(
                    f"✅ Данные водителя с ID {data['ID']} обновлены:\n\n{format_driver_data(data)}",
                    reply_markup=create_main_menu()
                )
                logger.info(f"Данные водителя {data.get('Водитель', 'Неизвестно')} с ID {data['ID']} обновлены")
            else:
                await message.answer(
                    f"⚠️ Не удалось обновить данные водителя с ID {data['ID']}. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Ошибка при обновлении данных водителя с ID {data.get('ID', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Обновление данных водителя отменено.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил обновление данных водителя")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении обновления водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обновлении данных водителя. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(lambda message: message.text == "📋 Запись")
async def add_transportation(message: types.Message, state: FSMContext):
    """
    Запрашивает данные перевозки и переходит в состояние добавления перевозки.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        await message.answer(
            "📝 Введите данные перевозки в следующем формате:\n"
            "Клиент: [название]\n"
            "Перевозчик: [название]\n"
            "Имя: [имя, если ИП]\n"
            "Телефон: [контакт]\n"
            "Водитель: [ФИО]\n"
            "Направление: [направление]\n"
            "Цена: [цена]\n"
            "Оплата: [оплата]\n"
            "Дата перевозки: [дд.мм.гггг]\n"
            "Пометка: [пометка]\n\n"
            "Пример:\n"
            "Клиент: Одуванчик\n"
            "Перевозчик: ИП Помидоров\n"
            "Имя: Помидоров Иван Иванович\n"
            "Телефон: +7 (123) 456-78-90\n"
            "Водитель: Иванов Иван Иванович\n"
            "Направление: Москва - Санкт-Петербург\n"
            "Цена: 50000\n"
            "Оплата: 45000\n"
            "Дата перевозки: 15.04.2025\n"
            "Пометка: Срочный заказ",
            reply_markup=create_record_submenu()
        )
        await state.set_state(Form.add_transportation)
        logger.info(f"Пользователь {message.from_user.id} выбрал запись перевозки")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка. Попробуйте снова.",
            reply_markup=create_main_menu()
        )

@dp.message(StateFilter(Form.add_transportation))
async def process_text_input(message: types.Message, state: FSMContext):
    """
    Обрабатывает данные о перевозке, добавляет запись в Excel и загружает на Яндекс.Диск.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        # Парсинг данных перевозки
        raw_data, data = parse_by_keywords(message.text, is_driver_data=False)
        if not data:
            await message.answer(
                "⚠️ Не удалось распознать данные. Пожалуйста, проверьте формат и попробуйте снова."
            )
            logger.warning(f"Не удалось распознать данные перевозки от пользователя {message.from_user.id}")
            return

        await db_manager.ensure_dbs_exist()

        # Проверка клиента
        client_name = data.get("Клиент", "")
        client_id = None
        if client_name:
            ws = db_manager.firms_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and client_name.lower() in row[1].lower():
                    client_id = row[0]
                    break
            if client_id is None:
                await message.answer(
                    f"⚠️ Клиент '{client_name}' не найден в базе. Пожалуйста, добавьте его через меню 'Добавить' -> 'Добавить фирму'.",
                    reply_markup=create_main_menu()
                )
                logger.warning(f"Клиент {client_name} не найден в базе")
                await state.clear()
                return
            existing_client = await db_manager.lookup_client(client_id)
            if not existing_client:
                await message.answer(
                    f"⚠️ Не удалось получить данные клиента '{client_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Не удалось получить данные клиента с ID {client_id}")
                await state.clear()
                return
            data["Фирма"] = existing_client["Полное название"]
            data["Клиент_ИНН"] = existing_client["ИНН"]
        else:
            await message.answer(
                "⚠️ Укажите клиента в формате 'Клиент: [название]'.",
                reply_markup=create_main_menu()
            )
            logger.warning(f"Клиент не указан в данных перевозки")
            await state.clear()
            return

        # Проверка перевозчика
        carrier_name = data.get("Перевозчик", "")
        carrier_id = None
        if carrier_name:
            ws = db_manager.carriers_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and carrier_name.lower() in row[1].lower():
                    carrier_id = row[0]
                    break
            if carrier_id is None:
                await message.answer(
                    f"⚠️ Перевозчик '{carrier_name}' не найден в базе. Пожалуйста, добавьте его через меню 'Добавить' -> 'Добавить перевозчика'.",
                    reply_markup=create_main_menu()
                )
                logger.warning(f"Перевозчик {carrier_name} не найден в базе")
                await state.clear()
                return
            existing_carrier = await db_manager.lookup_carrier(carrier_id)
            if not existing_carrier:
                await message.answer(
                    f"⚠️ Не удалось получить данные перевозчика '{carrier_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Не удалось получить данные перевозчика с ID {carrier_id}")
                await state.clear()
                return
            data["Перевозчик"] = existing_carrier["Полное название"]
            data["Перевозчик_ИНН"] = existing_carrier["ИНН"]
            data["Перевозчик_Контакт"] = existing_carrier["Контакт"]
        else:
            await message.answer(
                "⚠️ Укажите перевозчика в формате 'Перевозчик: [название]'.",
                reply_markup=create_main_menu()
            )
            logger.warning(f"Перевозчик не указан в данных перевозки")
            await state.clear()
            return

        # Проверка водителя
        driver_name = data.get("Водитель", "")
        driver_id = None
        if driver_name:
            ws = db_manager.drivers_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and driver_name.lower() in row[1].lower():
                    driver_id = row[0]
                    break
            if driver_id is None:
                await message.answer(
                    f"⚠️ Водитель '{driver_name}' не найден в базе. Пожалуйста, добавьте водителя через меню 'Добавить' -> 'Добавить водителя'.",
                    reply_markup=create_main_menu()
                )
                logger.warning(f"Водитель {driver_name} не найден в базе")
                await state.clear()
                return
            existing_driver = await db_manager.lookup_driver(driver_id)
            if not existing_driver:
                await message.answer(
                    f"⚠️ Не удалось получить данные водителя '{driver_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Не удалось получить данные водителя с ID {driver_id}")
                await state.clear()
                return
        else:
            await message.answer(
                "⚠️ Укажите водителя в формате 'Водитель: [ФИО]'.",
                reply_markup=create_main_menu()
            )
            logger.warning(f"Водитель не указан в данных перевозки")
            await state.clear()
            return

        # Обновление данных водителя в записи перевозки
        data.update({
            "Паспорт_серия_и_номер": existing_driver.get("Паспорт_серия_и_номер", ""),
            "Паспорт_место_выдачи": existing_driver.get("Паспорт_место_выдачи", ""),
            "Паспорт_дата_выдачи": existing_driver.get("Паспорт_дата_выдачи", ""),
            "Паспорт_код_подразделения": existing_driver.get("Паспорт_код_подразделения", ""),
            "ВУ_серия_и_номер": existing_driver.get("ВУ_серия_и_номер", ""),
            "В/У_дата_срок": existing_driver.get("В/У_дата_срок", ""),
            "Телефон": existing_driver.get("Телефон", ""),
            "Марка_машины_номер": existing_driver.get("Марка_машины_номер", ""),
            "Прицеп_номер": existing_driver.get("Прицеп_номер", ""),
            "Перевозчик": data["Перевозчик"],
            "Дата_рождения": existing_driver.get("Дата_рождения", ""),
            "Адрес_регистрации": existing_driver.get("Адрес_регистрации", ""),
            "Место_рождения": existing_driver.get("Место_рождения", ""),
            "Место_жительства": existing_driver.get("Место_жительства", ""),
            "Гражданство": existing_driver.get("Гражданство", "")
        })
        logger.debug(f"Данные водителя обновлены для записи перевозки: {data}")

        # Форматируем данные для подтверждения
        formatted_data = format_driver_data(data)
        await message.answer(
            f"Подтвердите данные перевозки:\n{formatted_data}\n\n"
            "Записать перевозку? (Да/Нет)",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
                ],
                resize_keyboard=True
            )
        )
        # Сохраняем ID для дальнейшего использования
        data['driver_id'] = driver_id
        data['client_id'] = client_id
        data['carrier_id'] = carrier_id
        await state.update_data(transportation_data=data)
        await state.set_state(Form.confirm_transportation)
        logger.info(f"Пользователь {message.from_user.id} ввёл данные перевозки")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обработке данных. Пожалуйста, попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(StateFilter(Form.confirm_transportation))
async def confirm_transportation(message: types.Message, state: FSMContext):
    """
    Подтверждает запись перевозки и сохраняет данные.

    Args:
        message (types.Message): Входящее сообщение.
        state (FSMContext): Контекст состояния.
    """
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_transportation, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("transportation_data", {})
        if not data:
            logger.error("transportation_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозки не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu()
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            excel_manager.ensure_files_exist()
            files = excel_manager.get_file_paths()
            success = True
            for file_type, path in files.items():
                if not excel_manager.add_record(path, data, file_type):
                    success = False
                    logger.error(f"Не удалось добавить запись в {path}")

            if success:
                record_info = "\n".join(
                    f"{key.replace('_', ' ')}: {value}" for key, value in data.items()
                    if value != "Не указано" and not key.endswith('_id')
                )
                await message.answer(
                    f"✅ Перевозка успешно записана:\n\n{record_info}",
                    reply_markup=create_main_menu()
                )
                logger.info(f"Перевозка успешно записана для водителя {data.get('Водитель', 'Неизвестно')}")

                await uploader.upload_files(list(files.values()), message)
            else:
                await message.answer(
                    "⚠️ Ошибка при записи перевозки. Данные сохранены локально. Попробуйте снова.",
                    reply_markup=create_main_menu()
                )
                logger.error(f"Ошибка при записи перевозки для водителя {data.get('Водитель', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Запись перевозки отменена.",
                reply_markup=create_main_menu()
            )
            logger.info(f"Пользователь {message.from_user.id} отменил запись перевозки")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении записи перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при записи перевозки. Попробуйте снова.",
            reply_markup=create_main_menu()
        )
        await state.clear()

# Запуск бота
async def main():
    """
    Запускает бота, инициализирует базы данных и начинает опрос сообщений.

    Raises:
        Exception: Если произошла ошибка при запуске.
    """
    try:
        logger.info("Запуск бота...")
        logger.info("Попытка инициализации баз данных...")
        await db_manager.ensure_dbs_exist()
        logger.info("Базы данных успешно инициализированы")
        logger.info("Проверка и создание Excel-файлов...")
        excel_manager.ensure_files_exist()
        logger.info("Excel-файлы успешно подготовлены")
        logger.info("Запуск polling для Telegram бота...")
        await dp.start_polling(bot)
        logger.info("Бот успешно запущен и работает")
    except Exception as e:
        logger.error(f"Не удалось инициализировать базы данных или запустить бота: {str(e)}")
        for path in [db_manager.drivers_db_path, db_manager.firms_db_path, db_manager.carriers_db_path]:
            if not os.path.exists(path):
                logger.warning(f"Файл {path} не найден, создание локальной версии")
                wb = Workbook()
                ws = wb.active
                if 'drivers' in path:
                    ws.append([
                        'ID', 'Водитель', 'Паспорт_серия_и_номер', 'Паспорт_место_выдачи',
                        'Паспорт_дата_выдачи', 'Паспорт_код_подразделения',
                        'ВУ_серия_и_номер', 'В/У_дата_срок', 'Телефон',
                        'Марка_машины_номер', 'Прицеп_номер', 'Перевозчик', 'Дата_рождения',
                        'Адрес_регистрации', 'Место_рождения', 'Место_жительства',
                        'Гражданство'
                    ])
                elif 'firms' in path:
                    ws.append(['ID', 'Краткое название', 'Полное название', 'ИНН'])
                elif 'carriers' in path:
                    ws.append(['ID', 'Краткое название', 'Полное название', 'ИНН', 'Контакт'])
                wb.save(path)
                logger.info(f"Создан локальный файл базы данных: {path}")
        await bot.send_message(
            chat_id=441196665,
            text="⚠️ Не удалось подключиться к Яндекс.Диску. Бот работает в локальном режиме."
        )
        raise

if __name__ == "__main__":
    try:
        logger.info("Запуск приложения...")
        asyncio.run(main())
    except Exception as e:
        logger.error(f"Критическая ошибка: {str(e)}\n{traceback.format_exc()}")
        sys.exit(1)

# Часть 7 завершена