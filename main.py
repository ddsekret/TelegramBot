import os
import logging
import asyncio
import yadisk
import traceback
import sys
import requests
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command, StateFilter
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from aiogram.client.session.aiohttp import AiohttpSession
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
from dotenv import load_dotenv
import pandas as pd
import re  # Добавлено

# Импорт функций парсинга из parser.py
from parser import (
    parse_driver_data,
    parse_car_data,
    parse_trailer_data,
    parse_carrier_data,
    parse_customer_data,
    normalize_data,
    parse_phone_numbers,
    parse_passport_issuing_authority,
    transliterate,
    validate_date,
)

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s",
    handlers=[
        logging.FileHandler("bot.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# Загрузка переменных окружения
load_dotenv()

# Токены
TELEGRAM_BOT_TOKEN = "8116572683:AAGxf2ttP-58uts18pRjTIy9cHX0LZfyUsU"
YANDEX_DISK_TOKEN = "y0__wgBELawpyAYkbY0IKiM5oQSstnwahr424ZzdNX_Y9dCWfPK-ac"
LOCAL_TEMP_DIR = "temp_files"

# Проверка токенов
if not TELEGRAM_BOT_TOKEN:
    logger.error("TELEGRAM_BOT_TOKEN не найден. Убедитесь, что он указан в .env файле.")
    sys.exit(1)

if not YANDEX_DISK_TOKEN:
    logger.error("YANDEX_DISK_TOKEN не найден. Убедитесь, что он указан в .env файле.")
    sys.exit(1)

# Логирование токенов для отладки
logger.debug(f"TELEGRAM_BOT_TOKEN: {TELEGRAM_BOT_TOKEN[:10]}...{TELEGRAM_BOT_TOKEN[-10:]}")
logger.debug(f"YANDEX_DISK_TOKEN: {YANDEX_DISK_TOKEN[:10]}...{YANDEX_DISK_TOKEN[-10:]}")

# Создание временной директории
os.makedirs(LOCAL_TEMP_DIR, exist_ok=True)
logger.info(f"Создана временная директория: {LOCAL_TEMP_DIR}")

# Глобальные константы
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DRIVERS_DB_PATH = os.path.join(BASE_DIR, "drivers_db.xlsx")
FIRMS_DB_PATH = os.path.join(BASE_DIR, "firms_db.xlsx")
CARRIERS_DB_PATH = os.path.join(BASE_DIR, "carriers_db.xlsx")
YANDEX_DRIVERS_PATH = "/TransportData/Database/drivers_db.xlsx"
YANDEX_FIRMS_PATH = "/TransportData/Database/firms_db.xlsx"
YANDEX_CARRIERS_PATH = "/TransportData/Database/carriers_db.xlsx"

# Инициализация YaDisk
y_disk = yadisk.AsyncYaDisk(token=YANDEX_DISK_TOKEN)
logger.info("YaDisk инициализирован")

# Инициализация бота и диспетчера
session = AiohttpSession(timeout=120)
bot = Bot(token=TELEGRAM_BOT_TOKEN, session=session)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
logger.info("Telegram Bot и Dispatcher инициализированы")


# Определение состояний
class Form(StatesGroup):
    add_driver = State()
    add_client = State()
    add_carrier = State()
    confirm_driver = State()
    confirm_client = State()
    confirm_carrier = State()
    add_transportation = State()
    confirm_transportation = State()


class DriverStates(StatesGroup):
    waiting_for_update_confirmation = State()
    waiting_for_add_confirmation = State()
    waiting_for_carrier_update_confirmation = State()


logger.info("Состояния Form и DriverStates инициализированы")


# Классы управления базами данных и Excel
class DatabaseManager:
    def __init__(self, drivers_db_path, firms_db_path, carriers_db_path):
        self.drivers_db_path = drivers_db_path
        self.firms_db_path = firms_db_path
        self.carriers_db_path = carriers_db_path
        self.drivers_wb = None
        self.firms_wb = None
        self.carriers_wb = None

    async def ensure_dbs_exist(self):
        try:
            for path, yandex_path, headers in [
                (
                    self.drivers_db_path,
                    YANDEX_DRIVERS_PATH,
                    [
                        "ID",
                        "Водитель",
                        "Паспорт_серия_и_номер",
                        "Паспорт_место_выдачи",
                        "Паспорт_дата_выдачи",
                        "Паспорт_код_подразделения",
                        "ВУ_серия_и_номер",
                        "В/У_дата_срок",
                        "Телефон",
                        "Автомобиль",
                        "Прицеп",
                        "Перевозчик",
                        "Дата_рождения",
                        "Адрес_регистрации",
                        "Место_рождения",
                        "Место_жительства",
                        "Гражданство",
                    ],
                ),
                (
                    self.firms_db_path,
                    YANDEX_FIRMS_PATH,
                    ["ID", "Краткое название", "Название", "ИНН"],
                ),
                (
                    self.carriers_db_path,
                    YANDEX_CARRIERS_PATH,
                    ["ID", "Краткое название", "Название", "ИНН", "Контакт"],
                ),
            ]:
                if not os.path.exists(path):
                    logger.info(f"Файл {path} не найден, попытка загрузки с Яндекс.Диска")
                    try:
                        await y_disk.download(yandex_path, path)
                        logger.info(f"Файл {path} успешно загружен с Яндекс.Диска")
                    except Exception as e:
                        logger.warning(f"Не удалось загрузить {path} с Яндекс.Диска: {str(e)}")
                        wb = Workbook()
                        ws = wb.active
                        ws.append(headers)
                        wb.save(path)
                        logger.info(f"Создан локальный файл базы данных: {path}")

            self.drivers_wb = load_workbook(self.drivers_db_path)
            self.firms_wb = load_workbook(self.firms_db_path)
            self.carriers_wb = load_workbook(self.carriers_db_path)
        except Exception as e:
            logger.error(f"Ошибка при инициализации баз данных: {str(e)}\n{traceback.format_exc()}")
            raise

    async def add_driver(self, data):
        try:
            ws = self.drivers_wb.active
            driver_id = ws.max_row
            row = [driver_id]
            fields = [
                "Водитель",
                "Паспорт_серия_и_номер",
                "Паспорт_место_выдачи",
                "Паспорт_дата_выдачи",
                "Паспорт_код_подразделения",
                "ВУ_серия_и_номер",
                "В/У_дата_срок",
                "Телефон",
                "Автомобиль",
                "Прицеп",
                "Перевозчик",
                "Дата_рождения",
                "Адрес_регистрации",
                "Место_рождения",
                "Место_жительства",
                "Гражданство",
            ]
            for field in fields:
                row.append(data.get(field, ""))
            ws.append(row)
            self.drivers_wb.save(self.drivers_db_path)
            logger.info(f"Водитель добавлен с ID {driver_id}")
            return driver_id
        except Exception as e:
            logger.error(f"Ошибка при добавлении водителя: {str(e)}\n{traceback.format_exc()}")
            return None

    async def update_driver(self, data):
        try:
            ws = self.drivers_wb.active
            driver_id = data.get("ID")
            for row in ws.iter_rows(min_row=2, max_col=1):
                if row[0].value == driver_id:
                    row_idx = row[0].row
                    fields = [
                        "Водитель",
                        "Паспорт_серия_и_номер",
                        "Паспорт_место_выдачи",
                        "Паспорт_дата_выдачи",
                        "Паспорт_код_подразделения",
                        "ВУ_серия_и_номер",
                        "В/У_дата_срок",
                        "Телефон",
                        "Автомобиль",
                        "Прицеп",
                        "Перевозчик",
                        "Дата_рождения",
                        "Адрес_регистрации",
                        "Место_рождения",
                        "Место_жительства",
                        "Гражданство",
                    ]
                    for idx, field in enumerate(fields, start=2):
                        ws.cell(row=row_idx, column=idx).value = data.get(field, "")
                    self.drivers_wb.save(self.drivers_db_path)
                    logger.info(f"Данные водителя с ID {driver_id} обновлены")
                    return True
            logger.error(f"Водитель с ID {driver_id} не найден")
            return False
        except Exception as e:
            logger.error(f"Ошибка при обновлении водителя: {str(e)}\n{traceback.format_exc()}")
            return False

    async def add_client(self, data):
        try:
            ws = self.firms_wb.active
            client_id = ws.max_row
            row = [
                client_id,
                data.get("Короткое название", ""),
                data.get("Название", ""),
                data.get("ИНН", ""),
            ]
            ws.append(row)
            self.firms_wb.save(self.firms_db_path)
            logger.info(f"Фирма добавлена с ID {client_id}")
            return client_id
        except Exception as e:
            logger.error(f"Ошибка при добавлении фирмы: {str(e)}\n{traceback.format_exc()}")
            return None

    async def add_carrier(self, data):
        try:
            ws = self.carriers_wb.active
            carrier_id = ws.max_row
            row = [
                carrier_id,
                data.get("Короткое название", ""),
                data.get("Перевозчик", ""),
                data.get("ИНН", ""),
                data.get("Контакт", ""),
            ]
            ws.append(row)
            self.carriers_wb.save(self.carriers_db_path)
            logger.info(f"Перевозчик добавлен с ID {carrier_id}")
            return carrier_id
        except Exception as e:
            logger.error(f"Ошибка при добавлении перевозчика: {str(e)}\n{traceback.format_exc()}")
            return None

    async def update_carrier(self, data):
        try:
            ws = self.carriers_wb.active
            carrier_id = data.get("ID")
            for row in ws.iter_rows(min_row=2, max_col=1):
                if row[0].value == carrier_id:
                    row_idx = row[0].row
                    ws.cell(row=row_idx, column=2).value = data.get("Короткое название", "")
                    ws.cell(row=row_idx, column=3).value = data.get("Перевозчик", "")
                    ws.cell(row=row_idx, column=4).value = data.get("ИНН", "")
                    ws.cell(row=row_idx, column=5).value = data.get("Контакт", "")
                    self.carriers_wb.save(self.carriers_db_path)
                    logger.info(f"Данные перевозчика с ID {carrier_id} обновлены")
                    return True
            logger.error(f"Перевозчик с ID {carrier_id} не найден")
            return False
        except Exception as e:
            logger.error(f"Ошибка при обновлении перевозчика: {str(e)}\n{traceback.format_exc()}")
            return False

    async def lookup_driver(self, driver_id):
        try:
            ws = self.drivers_wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == driver_id:
                    return {
                        "ID": row[0],
                        "Водитель": row[1],
                        "Паспорт_серия_и_номер": row[2],
                        "Паспорт_место_выдачи": row[3],
                        "Паспорт_дата_выдачи": row[4],
                        "Паспорт_код_подразделения": row[5],
                        "ВУ_серия_и_номер": row[6],
                        "В/У_дата_срок": row[7],
                        "Телефон": row[8],
                        "Автомобиль": row[9],
                        "Прицеп": row[10],
                        "Перевозчик": row[11],
                        "Дата_рождения": row[12],
                        "Адрес_регистрации": row[13],
                        "Место_рождения": row[14],
                        "Место_жительства": row[15],
                        "Гражданство": row[16],
                    }
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске водителя: {str(e)}\n{traceback.format_exc()}")
            return None

    async def lookup_client(self, client_id):
        try:
            ws = self.firms_wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == client_id:
                    return {
                        "ID": row[0],
                        "Краткое название": row[1],
                        "Название": row[2],
                        "ИНН": row[3],
                    }
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске клиента: {str(e)}\n{traceback.format_exc()}")
            return None

    async def lookup_carrier(self, carrier_name):
        try:
            ws = self.carriers_wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] and carrier_name.lower() in row[1].lower():
                    return {
                        "ID": row[0],
                        "Краткое название": row[1],
                        "Название": row[2],
                        "ИНН": row[3],
                        "Контакт": row[4],
                    }
            return None
        except Exception as e:
            logger.error(f"Ошибка при поиске перевозчика: {str(e)}\n{traceback.format_exc()}")
            return None


db_manager = DatabaseManager(DRIVERS_DB_PATH, FIRMS_DB_PATH, CARRIERS_DB_PATH)


class ExcelManager:
    def __init__(self):
        self.base_dir = BASE_DIR
        self.files = {
            "daily": os.path.join(self.base_dir, f"Перевозки_{datetime.now().strftime('%Y_%m_%d')}.xlsx"),
            "monthly": os.path.join(self.base_dir, f"Перевозки_{datetime.now().strftime('%Y_%m')}.xlsx"),
            "yearly": os.path.join(self.base_dir, f"Перевозки_{datetime.now().strftime('%Y')}.xlsx"),
        }

    def ensure_files_exist(self):
        for file_type, path in self.files.items():
            if not os.path.exists(path):
                wb = Workbook()
                ws = wb.active
                headers = [
                    "Фирма",
                    "Клиент_ИНН",
                    "Перевозчик",
                    "Перевозчик_ИНН",
                    "Перевозчик_Контакт",
                    "Водитель",
                    "Паспорт_серия_и_номер",
                    "Паспорт_место_выдачи",
                    "Паспорт_дата_выдачи",
                    "Паспорт_код_подразделения",
                    "ВУ_серия_и_номер",
                    "В/У_дата_срок",
                    "Телефон",
                    "Автомобиль",
                    "Прицеп",
                    "Направление",
                    "Цена",
                    "Оплата",
                    "Дата_перевозки",
                    "Пометка",
                    "Дата_рождения",
                    "Адрес_регистрации",
                    "Место_рождения",
                    "Место_жительства",
                    "Гражданство",
                ]
                ws.append(headers)
                wb.save(path)
                logger.info(f"Создан файл {path}")

    def get_file_paths(self):
        return self.files

    def add_record(self, path, data, file_type):
        try:
            wb = load_workbook(path)
            ws = wb.active
            row = []
            fields = [
                "Фирма",
                "Клиент_ИНН",
                "Перевозчик",
                "Перевозчик_ИНН",
                "Перевозчик_Контакт",
                "Водитель",
                "Паспорт_серия_и_номер",
                "Паспорт_место_выдачи",
                "Паспорт_дата_выдачи",
                "Паспорт_код_подразделения",
                "ВУ_серия_и_номер",
                "В/У_дата_срок",
                "Телефон",
                "Автомобиль",
                "Прицеп",
                "Направление",
                "Цена",
                "Оплата",
                "Дата_перевозки",
                "Пометка",
                "Дата_рождения",
                "Адрес_регистрации",
                "Место_рождения",
                "Место_жительства",
                "Гражданство",
            ]
            for field in fields:
                row.append(data.get(field, ""))
            ws.append(row)
            wb.save(path)
            logger.info(f"Запись добавлена в {path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при добавлении записи в {path}: {str(e)}\n{traceback.format_exc()}")
            return False


excel_manager = ExcelManager()


class YandexDiskUploader:
    async def upload_files(self, file_paths, message):
        try:
            for path in file_paths:
                filename = os.path.basename(path)
                yandex_path = f"/TransportData/Reports/{filename}"
                await y_disk.upload(path, yandex_path, overwrite=True)
                logger.info(f"Файл {filename} успешно загружен на Яндекс.Диск")
            await message.answer("✅ Файлы успешно загружены на Яндекс.Диск.")
        except Exception as e:
            logger.error(f"Ошибка при загрузке файлов на Яндекс.Диск: {str(e)}\n{traceback.format_exc()}")
            await message.answer("⚠️ Ошибка при загрузке файлов на Яндекс.Диск. Данные сохранены локально.")


uploader = YandexDiskUploader()


# Меню и хендлеры
def create_main_menu():
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📝 Добавить"), KeyboardButton(text="📋 Запись")],
            [KeyboardButton(text="🔍 Просмотр")],
        ],
        resize_keyboard=True,
    )
    return keyboard


def create_add_submenu():
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Добавить водителя")],
            [KeyboardButton(text="➕ Добавить фирму")],
            [KeyboardButton(text="➕ Добавить перевозчика")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
    )
    return keyboard


def create_record_submenu():
    keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="⬅️ Назад")]], resize_keyboard=True)
    return keyboard


def create_view_submenu():
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👤 Водители"), KeyboardButton(text="🏢 Фирмы")],
            [KeyboardButton(text="🚚 Перевозчики"), KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
    )
    return keyboard


def create_drivers_submenu():
    keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="⬅️ Назад")]], resize_keyboard=True)
    return keyboard


def create_firms_submenu():
    keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="⬅️ Назад")]], resize_keyboard=True)
    return keyboard


def create_carriers_submenu():
    keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="⬅️ Назад")]], resize_keyboard=True)
    return keyboard


def format_driver_data(data):
    result = []
    for key, value in data.items():
        if value is None or key.endswith("_id"):
            continue
        formatted_key = key.replace("_", " ")
        if key == "Прицеп_номер":
            formatted_key = "Прицеп"
        result.append(f"{formatted_key}: {value}")
    formatted_result = "\n".join(result)
    logger.debug(f"Formatted driver data: {formatted_result}")
    return formatted_result


@dp.message(Command("start"))
async def start_command(message: types.Message, state: FSMContext):
    try:
        await state.clear()
        await message.answer("👋 Добро пожаловать! Выберите действие:", reply_markup=create_main_menu())
        logger.info(f"Пользователь {message.from_user.id} запустил бота")
    except Exception as e:
        logger.error(f"Ошибка при выполнении команды /start: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "📝 Добавить")
async def add_menu(message: types.Message):
    try:
        await message.answer("Выберите, что хотите добавить:", reply_markup=create_add_submenu())
        logger.info(f"Пользователь {message.from_user.id} открыл меню добавления")
    except Exception as e:
        logger.error(f"Ошибка при открытии меню добавления: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "🔍 Просмотр")
async def view_menu(message: types.Message):
    try:
        await message.answer("Выберите, что хотите просмотреть:", reply_markup=create_view_submenu())
        logger.info(f"Пользователь {message.from_user.id} открыл меню просмотра")
    except Exception as e:
        logger.error(f"Ошибка при открытии меню просмотра: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "➕ Добавить водителя")
async def add_driver(message: types.Message, state: FSMContext):
    try:
        await message.answer(
            "📝 Введите данные водителя в следующем формате:\n"
            "Водитель: [ФИО]\n"
            "Паспорт: [серия и номер]\n"
            "Кем выдан: [место выдачи]\n"
            "Дата выдачи: [дд.мм.гггг]\n"
            "Код подразделения: [код]\n"
            "Водительское удостоверение: [серия и номер]\n"
            "Дата выдачи ВУ: [дд.мм.гггг]\n"
            "Телефон: [номер]\n"
            "Автомобиль: [марка, госномер]\n"
            "Прицеп: [марка, госномер]\n"
            "Перевозчик: [название]\n\n"
            "Пример:\n"
            "Водитель: Иванов Иван Иванович\n"
            "Паспорт: 1234 567890\n"
            "Кем выдан: УФМС России по г. Москве\n"
            "Дата выдачи: 01.01.2010\n"
            "Код подразделения: 123-456\n"
            "Водительское удостоверение: 12 34 567890\n"
            "Дата выдачи ВУ: 01.01.2015\n"
            "Телефон: +7 (123) 456-78-90\n"
            "Автомобиль: Volvo Р123РО123\n"
            "Прицеп: Schmitz АБ1234 12\n"
            "Перевозчик: ИП Иванов",
            reply_markup=create_record_submenu(),
        )
        await state.set_state(Form.add_driver)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление водителя")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "➕ Добавить фирму")
async def add_client(message: types.Message, state: FSMContext):
    try:
        await message.answer(
            "📝 Введите данные фирмы в следующем формате:\n"
            "Название: [название]\n"
            "ИНН: [ИНН]\n"
            "Короткое название: [короткое название, опционально]\n\n"
            "Пример:\n"
            "Название: ООО Ромашка\n"
            "ИНН: 1234567890\n"
            "Короткое название: Ромашка",
            reply_markup=create_record_submenu(),
        )
        await state.set_state(Form.add_client)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление фирмы")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "➕ Добавить перевозчика")
async def add_carrier(message: types.Message, state: FSMContext):
    try:
        await message.answer(
            "📝 Введите данные перевозчика в следующем формате:\n"
            "Перевозчик: [название]\n"
            "Имя: [имя, если ИП]\n"
            "Телефон: [контакт]\n"
            "ИНН: [ИНН]\n\n"
            "Пример:\n"
            "Перевозчик: ИП Атакишиев Маил Алиаббас Оглы\n"
            "Имя: Атакишиев Маил Алиаббас Оглы\n"
            "Телефон: +7 (921) 039-53-54\n"
            "ИНН: 519003116120",
            reply_markup=create_record_submenu(),
        )
        await state.set_state(Form.add_carrier)
        logger.info(f"Пользователь {message.from_user.id} выбрал добавление перевозчика")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(lambda message: message.text == "👤 Водители")
async def view_drivers(message: types.Message):
    try:
        await db_manager.ensure_dbs_exist()
        ws = db_manager.drivers_wb.active
        drivers_list = ""
        for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
            drivers_list += (
                f"ID: {row[0]}\n"
                f"Водитель: {row[1]}\n"
                f"Телефон: {row[8]}\n"
                f"Автомобиль: {row[9]}\n"
                f"Прицеп: {row[10]}\n"
                f"Перевозчик: {row[11]}\n"
                f"{'-' * 20}\n"
            )
        await message.answer(drivers_list, reply_markup=create_drivers_submenu())
        logger.info(f"Пользователь {message.from_user.id} запросил список водителей")
    except Exception as e:
        logger.error(f"Ошибка при просмотре водителей: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при загрузке списка водителей. Попробуйте снова.",
            reply_markup=create_drivers_submenu(),
        )


@dp.message(lambda message: message.text == "🏢 Фирмы")
async def view_firms(message: types.Message):
    try:
        await db_manager.ensure_dbs_exist()
        ws = db_manager.firms_wb.active
        firms_list = ""
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            firms_list += (
                f"ID: {row[0]}\n"
                f"Короткое название: {row[1]}\n"
                f"Название: {row[2]}\n"
                f"ИНН: {row[3]}\n"
                f"{'-' * 20}\n"
            )
        await message.answer(firms_list, reply_markup=create_firms_submenu())
        logger.info(f"Пользователь {message.from_user.id} запросил список фирм")
    except Exception as e:
        logger.error(f"Ошибка при просмотре фирм: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при загрузке списка фирм. Попробуйте снова.",
            reply_markup=create_firms_submenu(),
        )


@dp.message(lambda message: message.text == "🚚 Перевозчики")
async def view_carriers(message: types.Message):
    try:
        await db_manager.ensure_dbs_exist()
        ws = db_manager.carriers_wb.active
        carriers_list = ""
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            carriers_list += (
                f"ID: {row[0]}\n"
                f"Короткое название: {row[1]}\n"
                f"Название: {row[2]}\n"
                f"Контакт: {row[4]}\n"
                f"ИНН: {row[3]}\n"
                f"{'-' * 20}\n"
            )
        await message.answer(carriers_list, reply_markup=create_carriers_submenu())
        logger.info(f"Пользователь {message.from_user.id} запросил список перевозчиков")
    except Exception as e:
        logger.error(f"Ошибка при просмотре перевозчиков: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при загрузке списка перевозчиков. Попробуйте снова.",
            reply_markup=create_carriers_submenu(),
        )


@dp.message(lambda message: message.text == "⬅️ Назад")
async def go_back(message: types.Message, state: FSMContext):
    try:
        await state.clear()
        await message.answer("Вы вернулись в главное меню.", reply_markup=create_main_menu())
        logger.info(f"Пользователь {message.from_user.id} вернулся в главное меню")
    except Exception as e:
        logger.error(f"Ошибка при возвращении в главное меню: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(StateFilter(Form.add_driver))
async def process_add_driver(message: types.Message, state: FSMContext):
    try:
        await state.update_data(driver_data=message.text)
        user_data = await state.get_data()
        text = user_data.get("driver_data", "")
        if not text:
            await message.answer("Не удалось получить данные водителя. Пожалуйста, попробуйте снова.")
            logger.warning(f"Пустые данные водителя от пользователя {message.from_user.id}")
            return

        data = parse_driver_data(text)
        if "Водитель" not in data:
            logger.error("Поле 'Водитель' отсутствует в распарсенных данных")
            await message.answer("Не удалось определить ФИО водителя. Пожалуйста, проверьте формат данных.")
            return

        data = normalize_data(data)
        if not data:
            logger.error("Ошибка нормализации данных водителя")
            await message.answer("Ошибка обработки данных водителя. Пожалуйста, проверьте формат и попробуйте снова.")
            return

        await db_manager.ensure_dbs_exist()
        ws = db_manager.drivers_wb.active
        driver_id = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and data["Водитель"].lower() in row[1].lower():
                driver_id = row[0]
                break

        formatted_data = format_driver_data(data)
        if driver_id is not None:
            existing_driver = await db_manager.lookup_driver(driver_id)
            existing_formatted = format_driver_data(existing_driver)
            await message.answer(
                f"Водитель {data['Водитель']} уже есть в базе:\n\n"
                f"Текущие данные:\n{existing_formatted}\n\n"
                f"Новые данные:\n{formatted_data}\n\n"
                "Обновить данные? (Да/Нет)",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                    resize_keyboard=True,
                ),
            )
            data["ID"] = driver_id
            await state.update_data(driver_data=data)
            await state.set_state(DriverStates.waiting_for_update_confirmation)
            logger.info(f"Водитель {data['Водитель']} найден с ID {driver_id}, zapрошено обновление")
        else:
            driver_id = await db_manager.add_driver(data)
            if driver_id is not None:
                await message.answer(
                    f"Водитель {data['Водитель']} не найден в базе.\n\n"
                    f"Данные:\n{formatted_data}\n\n"
                    "Добавить водителя? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                        resize_keyboard=True,
                    ),
                )
                data["ID"] = driver_id
                await state.update_data(driver_data=data)
                await state.set_state(DriverStates.waiting_for_add_confirmation)
                logger.info(f"Водитель {data['Водитель']} не найден, запрошено добавление с ID {driver_id}")
            else:
                await message.answer(
                    "⚠️ Ошибка при добавлении водителя. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Ошибка при добавлении водителя: {data.get('Водитель', 'Неизвестно')}")
                await state.clear()

    except Exception as e:
        logger.error(f"Ошибка при обработке данных водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обработке данных. Пожалуйста, попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(Form.add_client))
async def process_add_client(message: types.Message, state: FSMContext):
    try:
        text = message.text.strip()
        client_data = parse_customer_data(text)
        if not client_data or "Название" not in client_data or "ИНН" not in client_data:
            logger.error("Не удалось извлечь название или ИНН фирмы")
            await message.answer(
                "Не удалось определить данные фирмы. Пожалуйста, укажите данные в формате:\n"
                "Название: [название]\n"
                "ИНН: [ИНН]\n"
                "Короткое название: [короткое название, опционально]\n\n"
                "Пример:\n"
                "Название: ООО Ромашка\n"
                "ИНН: 1234567890\n"
                "Короткое название: Ромашка",
                reply_markup=create_record_submenu(),
            )
            return

        client_data = normalize_data(client_data)

        full_name_lower = client_data["Название"].lower()
        inn = client_data["ИНН"]
        if (
            full_name_lower.startswith("ооо") or full_name_lower.startswith("оао") or full_name_lower.startswith("зао")
        ) and len(inn) != 10:
            await message.answer(
                "ИНН для ООО, ОАО или ЗАО должен содержать 10 цифр.",
                reply_markup=create_record_submenu(),
            )
            return
        elif full_name_lower.startswith("ип") and len(inn) != 12:
            await message.answer(
                "ИНН для ИП должен содержать 12 цифр.",
                reply_markup=create_record_submenu(),
            )
            return

        formatted_data = (
            f"Название: {client_data['Название']}\n"
            f"ИНН: {client_data['ИНН']}\n"
            f"Короткое название: {client_data.get('Короткое название', 'Не указано')}"
        )

        await message.answer(
            f"Вот как я разобрал данные фирмы:\n{formatted_data}\n\n" "Добавить фирму? (Да/Нет)",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                resize_keyboard=True,
            ),
        )
        ws = db_manager.firms_wb.active
        client_id = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and client_data["Короткое название"].lower() in row[1].lower():
                client_id = row[0]
                break
        if client_id is not None:
            client_data["ID"] = client_id
        await state.update_data(client_data=client_data)
        await state.set_state(Form.confirm_client)
        logger.info(f"Пользователь {message.from_user.id} ввёл данные фирмы: {client_data['Название']}")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_add_submenu())
        await state.clear()


@dp.message(StateFilter(Form.add_carrier))
async def process_add_carrier(message: types.Message, state: FSMContext):
    try:
        text = message.text.strip()
        logger.debug(f"Обработанный текст для парсинга: {text}")

        carrier_data = parse_carrier_data(text)
        if not carrier_data or "Перевозчик" not in carrier_data or "ИНН" not in carrier_data:
            logger.error("Не удалось извлечь название или ИНН перевозчика")
            await message.answer(
                "Не удалось определить данные перевозчика. Пожалуйста, укажите данные в формате:\n"
                "Перевозчик: [название]\n"
                "Имя: [имя, если ИП]\n"
                "Телефон: [контакт]\n"
                "ИНН: [ИНН]\n\n"
                "Пример:\n"
                "Перевозчик: ИП Атакишиев Маил Алиаббас Оглы\n"
                "Имя: Атакишиев Маил Алиаббас Оглы\n"
                "Телефон: +7 (921) 039-53-54\n"
                "ИНН: 519003116120",
                reply_markup=create_record_submenu(),
            )
            return

        carrier_data = normalize_data(carrier_data)

        carrier_name_lower = carrier_data["Перевозчик"].lower()
        inn = carrier_data["ИНН"]
        if carrier_name_lower.startswith("ип") and len(inn) != 12:
            await message.answer(
                "ИНН для ИП должен содержать 12 цифр.",
                reply_markup=create_record_submenu(),
            )
            return
        elif (
            carrier_name_lower.startswith("ооо")
            or carrier_name_lower.startswith("оао")
            or carrier_name_lower.startswith("зао")
        ) and len(inn) != 10:
            await message.answer(
                "ИНН для ООО, ОАО или ЗАО должен содержать 10 цифр.",
                reply_markup=create_record_submenu(),
            )
            return

        if "Контакт" not in carrier_data:
            await message.answer(
                "Некорректный формат телефона или телефон не указан. Укажите номер телефона, например: +7 (123) 456-78-90",
                reply_markup=create_record_submenu(),
            )
            return

        formatted_data = (
            f"Перевозчик: {carrier_data['Перевозчик']}\n"
            f"Имя: {carrier_data.get('Имя перевозчика', 'Не указано')}\n"
            f"Телефон: {carrier_data['Контакт']}\n"
            f"ИНН: {carrier_data['ИНН']}"
        )

        await db_manager.ensure_dbs_exist()
        ws = db_manager.carriers_wb.active
        logger.debug(f"Содержимое базы перевозчиков (carriers_db.xlsx):")
        for row in ws.iter_rows(min_row=2, values_only=True):
            logger.debug(f"Строка: {row}")
        carrier_id = None
        existing_carrier = None
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and carrier_data["Короткое название"].lower() in row[1].lower():
                carrier_id = row[0]
                break

        if carrier_id is not None:
            existing_carrier = await db_manager.lookup_carrier(carrier_data["Короткое название"])
            if existing_carrier:
                existing_formatted = (
                    f"Перевозчик: {existing_carrier['Название']}\n"
                    f"Имя: {existing_carrier.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {existing_carrier['Контакт']}\n"
                    f"ИНН: {existing_carrier['ИНН']}"
                )
                await message.answer(
                    f"Перевозчик {carrier_data['Короткое название']} уже есть в базе:\n\n"
                    f"Текущие данные:\n{existing_formatted}\n\n"
                    f"Новые данные:\n{formatted_data}\n\n"
                    "Обновить данные? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                        resize_keyboard=True,
                    ),
                )
                carrier_data["ID"] = carrier_id
                await state.update_data(carrier_data=carrier_data)
                await state.set_state(DriverStates.waiting_for_carrier_update_confirmation)
                logger.info(
                    f"Перевозчик {carrier_data['Короткое название']} найден с ID {carrier_id}, запрошено обновление"
                )
            else:
                logger.error(f"Не удалось получить данные перевозчика с ID {carrier_id}")
                await message.answer(
                    f"⚠️ Перевозчик {carrier_data['Короткое название']} уже существует в базе, но не удалось получить его данные.\n"
                    f"Новые данные:\n{formatted_data}\n\n"
                    "Обновить данные? (Да/Нет)",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                        resize_keyboard=True,
                    ),
                )
                carrier_data["ID"] = carrier_id
                await state.update_data(carrier_data=carrier_data)
                await state.set_state(DriverStates.waiting_for_carrier_update_confirmation)
        else:
            await message.answer(
                f"Вот как я разобрал данные перевозчика:\n{formatted_data}\n\n" "Добавить перевозчика? (Да/Нет)",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                    resize_keyboard=True,
                ),
            )
            await state.update_data(carrier_data=carrier_data)
            await state.set_state(Form.confirm_carrier)
            logger.info(f"Пользователь {message.from_user.id} ввёл данные перевозчика: {carrier_data['Перевозчик']}")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_add_submenu())
        await state.clear()


@dp.message(StateFilter(Form.confirm_client))
async def confirm_add_client(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_client, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        client_data = user_data.get("client_data", {})
        if not client_data:
            logger.error("client_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные фирмы не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            if "ID" in client_data:
                await message.answer(
                    "⚠️ Фирма уже существует в базе. Обновление данных фирмы не предусмотрено.",
                    reply_markup=create_main_menu(),
                )
                logger.info(
                    f"Пользователь {message.from_user.id} попытался добавить существующую фирму: {client_data['Название']}"
                )
            else:
                client_id = await db_manager.add_client(client_data)
                if client_id is not None:
                    formatted_data = (
                        f"Название: {client_data['Название']}\n"
                        f"ИНН: {client_data['ИНН']}\n"
                        f"Короткое название: {client_data.get('Короткое название', 'Не указано')}"
                    )
                    await message.answer(
                        f"✅ Фирма успешно добавлена с ID {client_id}:\n\n{formatted_data}",
                        reply_markup=create_main_menu(),
                    )
                    logger.info(f"Фирма {client_data['Название']} успешно добавлена с ID {client_id}")
                else:
                    await message.answer(
                        "⚠️ Ошибка при добавлении фирмы. Возможно, фирма с таким названием уже существует.",
                        reply_markup=create_main_menu(),
                    )
                    logger.error(f"Ошибка при добавлении фирмы {client_data.get('Название', 'Неизвестно')}")
        else:
            await message.answer("⏭ Добавление фирмы отменено.", reply_markup=create_main_menu())
            logger.info(f"Пользователь {message.from_user.id} отменил добавление фирмы")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления фирмы: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении фирмы. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(Form.confirm_carrier))
async def confirm_add_carrier(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_carrier, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        carrier_data = user_data.get("carrier_data", {})
        carrier_name = carrier_data.get("Короткое название", "Неизвестно")
        if not carrier_data:
            logger.error("carrier_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозчика не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            existing_carrier = await db_manager.lookup_carrier(carrier_name)
            if existing_carrier:
                await message.answer(
                    f"⚠️ Перевозчик {carrier_name} уже существует в базе:\n"
                    f"Перевозчик: {existing_carrier['Название']}\n"
                    f"Имя: {existing_carrier.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {existing_carrier['Контакт']}\n"
                    f"ИНН: {existing_carrier['ИНН']}\n\n"
                    "Чтобы обновить данные, повторите ввод с изменёнными данными.",
                    reply_markup=create_main_menu(),
                )
                logger.info(f"Попытка добавить уже существующего перевозчика: {carrier_name}")
            else:
                carrier_id = await db_manager.add_carrier(carrier_data)
                if carrier_id is not None:
                    formatted_data = (
                        f"Перевозчик: {carrier_data['Перевозчик']}\n"
                        f"Имя: {carrier_data.get('Имя перевозчика', 'Не указано')}\n"
                        f"Телефон: {carrier_data['Контакт']}\n"
                        f"ИНН: {carrier_data['ИНН']}"
                    )
                    await message.answer(
                        f"✅ Перевозчик успешно добавлен с ID {carrier_id}:\n\n{formatted_data}",
                        reply_markup=create_main_menu(),
                    )
                    logger.info(f"Перевозчик {carrier_data['Перевозчик']} успешно добавлен с ID {carrier_id}")
                else:
                    await message.answer(
                        "⚠️ Ошибка при добавлении перевозчика. Возможно, перевозчик с таким названием уже существует.",
                        reply_markup=create_main_menu(),
                    )
                    logger.error(f"Ошибка при добавлении перевозчика {carrier_data.get('Перевозчик', 'Неизвестно')}")
        else:
            await message.answer("⏭ Добавление перевозчика отменено.", reply_markup=create_main_menu())
            logger.info(f"Пользователь {message.from_user.id} отменил добавление перевозчика")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении перевозчика. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(DriverStates.waiting_for_carrier_update_confirmation))
async def confirm_update_carrier(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_update_carrier, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        carrier_data = user_data.get("carrier_data", {})
        carrier_name = carrier_data.get("Короткое название", "Неизвестно")
        if not carrier_data:
            logger.error("carrier_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозчика не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if "ID" not in carrier_data:
            logger.error("ID перевозчика отсутствует в данных")
            await message.answer(
                "⚠️ ID перевозчика не найден. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            success = await db_manager.update_carrier(carrier_data)
            if success:
                formatted_data = (
                    f"Перевозчик: {carrier_data['Перевозчик']}\n"
                    f"Имя: {carrier_data.get('Имя перевозчика', 'Не указано')}\n"
                    f"Телефон: {carrier_data['Контакт']}\n"
                    f"ИНН: {carrier_data['ИНН']}"
                )
                await message.answer(
                    f"✅ Данные перевозчика с ID {carrier_data['ID']} обновлены:\n\n{formatted_data}",
                    reply_markup=create_main_menu(),
                )
                logger.info(f"Данные перевозчика {carrier_data['Перевозчик']} с ID {carrier_data['ID']} обновлены")
            else:
                await message.answer(
                    f"⚠️ Не удалось обновить данные перевозчика {carrier_name}. Возможно, перевозчик с ID {carrier_data['ID']} не найден.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Ошибка при обновлении данных перевозчика с ID {carrier_data.get('ID', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Обновление данных перевозчика отменено.",
                reply_markup=create_main_menu(),
            )
            logger.info(f"Пользователь {message.from_user.id} отменил обновление данных перевозчика")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении обновления перевозчика: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обновлении данных перевозчика. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(DriverStates.waiting_for_add_confirmation))
async def confirm_add_driver(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_add_driver, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("driver_data", {})
        if not data:
            logger.error("driver_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные водителя не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            await message.answer(
                f"✅ Водитель успешно добавлен с ID {data['ID']}:\n\n{format_driver_data(data)}",
                reply_markup=create_main_menu(),
            )
            logger.info(f"Водитель {data.get('Водитель', 'Неизвестно')} успешно подтверждён с ID {data['ID']}")
        else:
            await message.answer("⏭ Добавление водителя отменено.", reply_markup=create_main_menu())
            logger.info(f"Пользователь {message.from_user.id} отменил добавление водителя")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении водителя. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()

        if message.text.lower() == "да":
            await message.answer(
                f"✅ Водитель успешно добавлен с ID {data['ID']}:\n\n{format_driver_data(data)}",
                reply_markup=create_main_menu(),
            )
            logger.info(f"Водитель {data.get('Водитель', 'Неизвестно')} успешно подтверждён с ID {data['ID']}")
        else:
            await message.answer("⏭ Добавление водителя отменено.", reply_markup=create_main_menu())
            logger.info(f"Пользователь {message.from_user.id} отменил добавление водителя")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении добавления водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при добавлении водителя. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(DriverStates.waiting_for_update_confirmation))
async def confirm_update_driver(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_update_driver, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("driver_data", {})
        if not data:
            logger.error("driver_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные водителя не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            success = await db_manager.update_driver(data)
            if success:
                await message.answer(
                    f"✅ Данные водителя с ID {data['ID']} обновлены:\n\n{format_driver_data(data)}",
                    reply_markup=create_main_menu(),
                )
                logger.info(f"Данные водителя {data.get('Водитель', 'Неизвестно')} с ID {data['ID']} обновлены")
            else:
                await message.answer(
                    f"⚠️ Не удалось обновить данные водителя с ID {data['ID']}. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Ошибка при обновлении данных водителя с ID {data.get('ID', 'Неизвестно')}")
        else:
            await message.answer(
                "⏭ Обновление данных водителя отменено.",
                reply_markup=create_main_menu(),
            )
            logger.info(f"Пользователь {message.from_user.id} отменил обновление данных водителя")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении обновления водителя: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обновлении данных водителя. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(lambda message: message.text == "📋 Запись")
async def add_transportation(message: types.Message, state: FSMContext):
    try:
        await message.answer(
            "📝 Введите данные перевозки в следующем формате:\n"
            "Клиент: [название]\n"
            "Перевозчик: [название]\n"
            "Имя: [имя, если ИП]\n"
            "Телефон: [контакт]\n"
            "Водитель: [ФИО]\n"
            "Направление: [направление]\n"
            "Цена: [цена]\n"
            "Оплата: [оплата]\n"
            "Дата перевозки: [дд.мм.гггг]\n"
            "Пометка: [пометка]\n\n"
            "Пример:\n"
            "Клиент: ООО Ромашка\n"
            "Перевозчик: ИП Помидоров\n"
            "Имя: Помидоров Иван Иванович\n"
            "Телефон: +7 (123) 456-78-90\n"
            "Водитель: Иванов Иван Иванович\n"
            "Направление: Москва - Санкт-Петербург\n"
            "Цена: 50000\n"
            "Оплата: 45000\n"
            "Дата перевозки: 15.04.2025\n"
            "Пометка: Срочный заказ",
            reply_markup=create_record_submenu(),
        )
        await state.set_state(Form.add_transportation)
        logger.info(f"Пользователь {message.from_user.id} выбрал запись перевозки")
    except Exception as e:
        logger.error(f"Ошибка при запросе данных перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer("⚠️ Произошла ошибка. Попробуйте снова.", reply_markup=create_main_menu())


@dp.message(StateFilter(Form.add_transportation))
async def process_text_input(message: types.Message, state: FSMContext):
    try:
        text = message.text.strip()
        data = {}
        lines = text.split("\n")

        client_lines = [
            line for line in lines if line.lower().startswith("клиент") or line.lower().startswith("заказчик")
        ]
        if client_lines:
            client_text = "\n".join(client_lines)
            client_data = parse_customer_data(client_text)
            if client_data and "Название" in client_data:
                data.update({"Фирма": client_data["Название"], "Клиент_ИНН": client_data["ИНН"]})
            else:
                await message.answer(
                    "⚠️ Не удалось распознать данные клиента. Укажите в формате 'Клиент: [название]'.",
                    reply_markup=create_record_submenu(),
                )
                logger.warning(f"Не удалось распознать данные клиента от пользователя {message.from_user.id}")
                return

        carrier_lines = [
            line for line in lines if line.lower().startswith("перевозчик") or line.lower().startswith("превозчик")
        ]
        if carrier_lines:
            carrier_text = "\n".join(carrier_lines)
            carrier_data = parse_carrier_data(carrier_text)
            if carrier_data and "Перевозчик" in carrier_data:
                data.update(
                    {
                        "Перевозчик": carrier_data["Перевозчик"],
                        "Перевозчик_ИНН": carrier_data["ИНН"],
                        "Перевозчик_Контакт": carrier_data.get("Контакт", ""),
                    }
                )
            else:
                await message.answer(
                    "⚠️ Не удалось распознать данные перевозчика. Укажите в формате 'Перевозчик: [название]'.",
                    reply_markup=create_record_submenu(),
                )
                logger.warning(f"Не удалось распознать данные перевозчика от пользователя {message.from_user.id}")
                return

        for line in lines:
            line = line.strip()
            if line.lower().startswith("водитель"):
                match = re.match(r"(?:Водитель)\s*[:\-\s]*(.+)", line, re.IGNORECASE)
                if match:
                    data["Водитель"] = match.group(1).strip()
            elif line.lower().startswith("направление"):
                match = re.match(r"(?:Направление)\s*[:\-\s]*(.+)", line, re.IGNORECASE)
                if match:
                    data["Направление"] = match.group(1).strip()
            elif line.lower().startswith("цена"):
                match = re.match(r"(?:Цена)\s*[:\-\s]*(\d+)", line, re.IGNORECASE)
                if match:
                    data["Цена"] = int(match.group(1))
            elif line.lower().startswith("оплата"):
                match = re.match(r"(?:Оплата)\s*[:\-\s]*(\d+)", line, re.IGNORECASE)
                if match:
                    data["Оплата"] = int(match.group(1))
            elif line.lower().startswith("дата перевозки"):
                match = re.match(
                    r"(?:Дата\s*перевозки)\s*[:\-\s]*(\d{2}\.\d{2}\.\d{4})",
                    line,
                    re.IGNORECASE,
                )
                if match and validate_date(match.group(1)):
                    data["Дата_перевозки"] = match.group(1)
            elif line.lower().startswith("пометка"):
                match = re.match(r"(?:Пометка)\s*[:\-\s]*(.+)", line, re.IGNORECASE)
                if match:
                    data["Пометка"] = match.group(1).strip()

        if not data.get("Водитель"):
            await message.answer(
                "⚠️ Укажите водителя в формате 'Водитель: [ФИО]'.",
                reply_markup=create_record_submenu(),
            )
            logger.warning(f"Водитель не указан в данных перевозки")
            return

        if not data.get("Фирма"):
            await message.answer(
                "⚠️ Укажите клиента в формате 'Клиент: [название]'.",
                reply_markup=create_record_submenu(),
            )
            logger.warning(f"Клиент не указан в данных перевозки")
            return

        if not data.get("Перевозчик"):
            await message.answer(
                "⚠️ Укажите перевозчика в формате 'Перевозчик: [название]'.",
                reply_markup=create_record_submenu(),
            )
            logger.warning(f"Перевозчик не указан в данных перевозки")
            return

        data = normalize_data(data)
        if not data:
            logger.error("Ошибка нормализации данных перевозки")
            await message.answer(
                "⚠️ Ошибка обработки данных перевозки. Пожалуйста, проверьте формат и попробуйте снова.",
                reply_markup=create_record_submenu(),
            )
            return

        await db_manager.ensure_dbs_exist()

        client_name = data.get("Фирма", "")
        client_id = None
        if client_name:
            ws = db_manager.firms_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and client_name.lower() in row[1].lower():
                    client_id = row[0]
                    break
            if client_id is None:
                await message.answer(
                    f"⚠️ Клиент '{client_name}' не найден в базе. Пожалуйста, добавьте его через меню 'Добавить' -> 'Добавить фирму'.",
                    reply_markup=create_main_menu(),
                )
                logger.warning(f"Клиент {client_name} не найден в базе")
                await state.clear()
                return
            existing_client = await db_manager.lookup_client(client_id)
            if not existing_client:
                await message.answer(
                    f"⚠️ Не удалось получить данные клиента '{client_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Не удалось получить данные клиента с ID {client_id}")
                await state.clear()
                return
            data["Фирма"] = existing_client["Название"]
            data["Клиент_ИНН"] = existing_client["ИНН"]
        else:
            await message.answer(
                "⚠️ Укажите клиента в формате 'Клиент: [название]'.",
                reply_markup=create_main_menu(),
            )
            logger.warning(f"Клиент не указан в данных перевозки")
            await state.clear()
            return

        carrier_name = data.get("Перевозчик", "")
        carrier_id = None
        if carrier_name:
            ws = db_manager.carriers_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and carrier_name.lower() in row[1].lower():
                    carrier_id = row[0]
                    break
            if carrier_id is None:
                await message.answer(
                    f"⚠️ Перевозчик '{carrier_name}' не найден в базе. Пожалуйста, добавьте его через меню 'Добавить' -> 'Добавить перевозчика'.",
                    reply_markup=create_main_menu(),
                )
                logger.warning(f"Перевозчик {carrier_name} не найден в базе")
                await state.clear()
                return
            existing_carrier = await db_manager.lookup_carrier(carrier_id)
            if not existing_carrier:
                await message.answer(
                    f"⚠️ Не удалось получить данные перевозчика '{carrier_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Не удалось получить данные перевозчика с ID {carrier_id}")
                await state.clear()
                return
            data["Перевозчик"] = existing_carrier["Название"]
            data["Перевозчик_ИНН"] = existing_carrier["ИНН"]
            data["Перевозчик_Контакт"] = existing_carrier["Контакт"]
        else:
            await message.answer(
                "⚠️ Укажите перевозчика в формате 'Перевозчик: [название]'.",
                reply_markup=create_main_menu(),
            )
            logger.warning(f"Перевозчик не указан в данных перевозки")
            await state.clear()
            return

        driver_name = data.get("Водитель", "")
        driver_id = None
        if driver_name:
            ws = db_manager.drivers_wb.active
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] and driver_name.lower() in row[1].lower():
                    driver_id = row[0]
                    break
            if driver_id is None:
                await message.answer(
                    f"⚠️ Водитель '{driver_name}' не найден в базе. Пожалуйста, добавьте водителя через меню 'Добавить' -> 'Добавить водителя'.",
                    reply_markup=create_main_menu(),
                )
                logger.warning(f"Водитель {driver_name} не найден в базе")
                await state.clear()
                return
            existing_driver = await db_manager.lookup_driver(driver_id)
            if not existing_driver:
                await message.answer(
                    f"⚠️ Не удалось получить данные водителя '{driver_name}' из базы. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Не удалось получить данные водителя с ID {driver_id}")
                await state.clear()
                return
        else:
            await message.answer(
                "⚠️ Укажите водителя в формате 'Водитель: [ФИО]'.",
                reply_markup=create_main_menu(),
            )
            logger.warning(f"Водитель не указан в данных перевозки")
            await state.clear()
            return

        data.update(
            {
                "Паспорт_серия_и_номер": existing_driver.get("Паспорт_серия_и_номер", ""),
                "Паспорт_место_выдачи": existing_driver.get("Паспорт_место_выдачи", ""),
                "Паспорт_дата_выдачи": existing_driver.get("Паспорт_дата_выдачи", ""),
                "Паспорт_код_подразделения": existing_driver.get("Паспорт_код_подразделения", ""),
                "ВУ_серия_и_номер": existing_driver.get("ВУ_серия_и_номер", ""),
                "В/У_дата_срок": existing_driver.get("В/У_дата_срок", ""),
                "Телефон": existing_driver.get("Телефон", ""),
                "Автомобиль": existing_driver.get("Автомобиль", ""),
                "Прицеп": existing_driver.get("Прицеп", ""),
                "Перевозчик": data["Перевозчик"],
                "Дата_рождения": existing_driver.get("Дата_рождения", ""),
                "Адрес_регистрации": existing_driver.get("Адрес_регистрации", ""),
                "Место_рождения": existing_driver.get("Место_рождения", ""),
                "Место_жительства": existing_driver.get("Место_жительства", ""),
                "Гражданство": existing_driver.get("Гражданство", ""),
            }
        )
        logger.debug(f"Данные водителя обновлены для записи перевозки: {data}")

        formatted_data = format_driver_data(data)
        await message.answer(
            f"Подтвердите данные перевозки:\n{formatted_data}\n\n" "Записать перевозку? (Да/Нет)",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
                resize_keyboard=True,
            ),
        )
        data["driver_id"] = driver_id
        data["client_id"] = client_id
        data["carrier_id"] = carrier_id
        await state.update_data(transportation_data=data)
        await state.set_state(Form.confirm_transportation)
        logger.info(f"Пользователь {message.from_user.id} ввёл данные перевозки")
    except Exception as e:
        logger.error(f"Ошибка при обработке данных перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при обработке данных. Пожалуйста, попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


@dp.message(StateFilter(Form.confirm_transportation))
async def confirm_transportation(message: types.Message, state: FSMContext):
    try:
        current_state = await state.get_state()
        logger.debug(f"Вызван confirm_transportation, текущее состояние: {current_state}")
        user_data = await state.get_data()
        logger.debug(f"Данные состояния: {user_data}")
        data = user_data.get("transportation_data", {})
        if not data:
            logger.error("transportation_data отсутствует в состоянии")
            await message.answer(
                "⚠️ Данные перевозки не найдены. Пожалуйста, повторите ввод.",
                reply_markup=create_main_menu(),
            )
            await state.clear()
            return

        if message.text.lower() == "да":
            excel_manager.ensure_files_exist()
            files = excel_manager.get_file_paths()
            success = True
            for file_type, path in files.items():
                if not excel_manager.add_record(path, data, file_type):
                    success = False
                    logger.error(f"Не удалось добавить запись в {path}")

            if success:
                record_info = "\n".join(
                    f"{key.replace('_', ' ')}: {value}"
                    for key, value in data.items()
                    if value != "Не указано" and not key.endswith("_id")
                )
                await message.answer(
                    f"✅ Перевозка успешно записана:\n\n{record_info}",
                    reply_markup=create_main_menu(),
                )
                logger.info(f"Перевозка успешно записана для водителя {data.get('Водитель', 'Неизвестно')}")

                await uploader.upload_files(list(files.values()), message)
            else:
                await message.answer(
                    "⚠️ Ошибка при записи перевозки. Данные сохранены локально. Попробуйте снова.",
                    reply_markup=create_main_menu(),
                )
                logger.error(f"Ошибка при записи перевозки для водителя {data.get('Водитель', 'Неизвестно')}")
        else:
            await message.answer("⏭ Запись перевозки отменена.", reply_markup=create_main_menu())
            logger.info(f"Пользователь {message.from_user.id} отменил запись перевозки")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при подтверждении записи перевозки: {str(e)}\n{traceback.format_exc()}")
        await message.answer(
            "⚠️ Произошла ошибка при записи перевозки. Попробуйте снова.",
            reply_markup=create_main_menu(),
        )
        await state.clear()


# Запуск бота
async def main():
    """
    Запускает бота, инициализирует базы данных и начинает опрос сообщений.

    Raises:
        Exception: Если произошла ошибка при запуске.
    """
    try:
        logger.info("Запуск бота...")
        logger.info("Попытка инициализации баз данных...")
        await db_manager.ensure_dbs_exist()
        logger.info("Базы данных успешно инициализированы")
        logger.info("Проверка и создание Excel-файлов...")
        excel_manager.ensure_files_exist()
        logger.info("Excel-файлы успешно подготовлены")
        logger.info("Запуск polling для Telegram бота...")
        await dp.start_polling(bot)
        logger.info("Бот успешно запущен и работает")
    except Exception as e:
        logger.error(f"Не удалось инициализировать базы данных или запустить бота: {str(e)}")
        for path in [
            db_manager.drivers_db_path,
            db_manager.firms_db_path,
            db_manager.carriers_db_path,
        ]:
            if not os.path.exists(path):
                logger.warning(f"Файл {path} не найден, создание локальной версии")
                wb = Workbook()
                ws = wb.active
                if "drivers" in path:
                    ws.append(
                        [
                            "ID",
                            "Водитель",
                            "Паспорт_серия_и_номер",
                            "Паспорт_место_выдачи",
                            "Паспорт_дата_выдачи",
                            "Паспорт_код_подразделения",
                            "ВУ_серия_и_номер",
                            "В/У_дата_срок",
                            "Телефон",
                            "Автомобиль",
                            "Прицеп",
                            "Перевозчик",
                            "Дата_рождения",
                            "Адрес_регистрации",
                            "Место_рождения",
                            "Место_жительства",
                            "Гражданство",
                        ]
                    )
                elif "firms" in path:
                    ws.append(["ID", "Краткое название", "Название", "ИНН"])
                elif "carriers" in path:
                    ws.append(["ID", "Краткое название", "Название", "ИНН", "Контакт"])
                wb.save(path)
                logger.info(f"Создан локальный файл базы данных: {path}")
        await bot.send_message(
            chat_id=441196665,  # Замените на ваш chat_id для уведомлений
            text="⚠️ Не удалось подключиться к Яндекс.Диску. Бот работает в локальном режиме.",
        )
        raise


if __name__ == "__main__":
    try:
        logger.info("Запуск приложения...")
        asyncio.run(main())
    except Exception as e:
        logger.error(f"Критическая ошибка: {str(e)}\n{traceback.format_exc()}")
        sys.exit(1)
